# app.py
# Referee Allocator (MVP) — Admin + Referee Portal + Offers + Blackouts + Printable PDFs

import os
import sqlite3
import secrets
import smtplib
import streamlit.components.v1 as components
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from pathlib import Path
from datetime import datetime, date, timedelta, timezone
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

import pandas as pd
import streamlit as st
from dateutil import parser as dtparser
from streamlit_autorefresh import st_autorefresh

# PDF (ReportLab)
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.pdfgen import canvas
from reportlab.lib.units import mm

def preserve_scroll(scroll_key: str = "refalloc_admin_scroll"):
    """
    Persists the page scroll position (window.scrollY) in localStorage and
    restores it after every Streamlit rerun (including st_autorefresh).

    - Uses one global scroll listener
    - Updates the active storage KEY every rerun (supports per-tab/per-date keys)
    """
    components.html(
        f"""
        <script>
        (function() {{
          // Always update the active key (so one listener supports many dates/tabs)
          window.__refallocScrollKey = "{scroll_key}";

          // Install scroll listener only once per page load
          if (!window.__refallocScrollInstalled) {{
            window.__refallocScrollInstalled = true;

            let ticking = false;
            window.addEventListener("scroll", function() {{
              if (!ticking) {{
                window.requestAnimationFrame(function() {{
                  try {{
                    const KEY = window.__refallocScrollKey || "{scroll_key}";
                    localStorage.setItem(KEY, String(window.scrollY || 0));
                  }} catch (e) {{}}
                  ticking = false;
                }});
                ticking = true;
              }}
            }}, {{ passive: true }});
          }}

          function restore() {{
            let y = 0;
            try {{
              const KEY = window.__refallocScrollKey || "{scroll_key}";
              y = parseInt(localStorage.getItem(KEY) || "0", 10) || 0;
            }} catch (e) {{}}

            const maxY = Math.max(0, document.body.scrollHeight - window.innerHeight);
            if (y > maxY) y = maxY;

            window.scrollTo(0, y);
          }}

          window.setTimeout(restore, 0);
          window.setTimeout(restore, 80);
          window.setTimeout(restore, 200);
        }})();
        </script>
        """,
        height=0,
        width=0,
    )

# ============================
# Admin auth helpers
# ============================

SUPER_ADMIN_EMAIL = "landon737@gmail.com"

def is_super_admin_logged_in() -> bool:
    return (st.session_state.get("admin_email", "").strip().lower() == SUPER_ADMIN_EMAIL)

def is_super_admin_email(email: str) -> bool:
    return (email or "").strip().lower() == SUPER_ADMIN_EMAIL


# ============================================================
# CONFIG
# ============================================================
BASE_DIR = Path(__file__).resolve().parent

DB_PATH = os.getenv("DB_PATH", str(BASE_DIR / "league.db"))
Path(DB_PATH).expanduser().parent.mkdir(parents=True, exist_ok=True)

REF_PORTAL_ENABLED = os.getenv("REF_PORTAL_ENABLED", "false").lower() == "true"
DEBUG_BANNER = os.getenv("DEBUG_BANNER", "false").lower() == "true"

st.set_page_config(page_title="Referee Allocator (MVP)", layout="wide")

if DEBUG_BANNER:
    st.warning("DEBUG: App reached top of script ✅")
    st.write("DEBUG: query_params =", dict(st.query_params))
    st.write("DEBUG: session_state keys =", list(st.session_state.keys()))
    st.markdown("---")


# ============================================================
# ⚠️ CORE HELPERS — DO NOT MOVE BELOW THIS LINE
# ============================================================

# ============================================================
# Small utilities
# ============================================================
def game_local_date(game_row) -> date:
    """
    Returns the local calendar date for a game row.
    Expects game_row to have 'start_dt' (ISO string).
    """
    s = (game_row["start_dt"] or "").strip()

    # Accept both "YYYY-MM-DD HH:MM" and ISO "YYYY-MM-DDTHH:MM:SS"
    s = s.replace(" ", "T")

    dt = dtparser.isoparse(s)
    return dt.date()

def parse_csv_date_strict(d_raw: str) -> date:
    """
    Parse a date string from CSV with ZERO ambiguity.
    Accepts:
      - YYYY-MM-DD   (preferred)
      - DD/MM/YYYY   (common NZ)
      - DD-MM-YYYY
    """
    s = (d_raw or "").strip()
    if not s:
        raise ValueError("Empty date")

    # If pandas gave us "2026-02-04 00:00:00"
    s = s.split(" ")[0]

    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass

    raise ValueError(f"Unrecognised date format: '{d_raw}' (expected YYYY-MM-DD or DD/MM/YYYY)")

def referee_has_blackout(ref_id: int, d: date) -> bool:
    """
    Returns True if the referee has a blackout on the given date.
    """
    conn = db()
    row = conn.execute(
        """
        SELECT 1
        FROM blackouts
        WHERE referee_id=? AND blackout_date=?
        LIMIT 1
        """,
        (ref_id, d.isoformat()),
    ).fetchone()
    conn.close()
    return bool(row)


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def status_badge(text: str, bg: str, fg: str = "white"):
    st.markdown(
        f"""
        <div style="
            display:inline-block;
            padding:6px 10px;
            border-radius:8px;
            background:{bg};
            color:{fg};
            font-weight:700;
            font-size:14px;
        ">
        {text}
        </div>
        """,
        unsafe_allow_html=True,
    )


def _time_12h(dt: datetime) -> str:
    return dt.strftime("%I:%M %p").lstrip("0")

def invalidate_ladder() -> None:
    """
    Force Streamlit to recompute ladder tables after saving a result.
    Works whether you use @st.cache_data or not.
    """
    try:
        st.cache_data.clear()
    except Exception:
        pass

    st.session_state["ladder_nonce"] = str(datetime.now(timezone.utc).timestamp())


# ============================================================
# DB
# ============================================================
def db():
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON;")
    conn.execute("PRAGMA journal_mode = WAL;")
    conn.execute("PRAGMA synchronous = NORMAL;")
    conn.execute("PRAGMA busy_timeout = 5000;")
    return conn

# ============================================================
# Backup / Restore (SQLite online backup API)
# ============================================================

BACKUPS_DIR = Path(DB_PATH).expanduser().parent / "backups"
BACKUPS_DIR.mkdir(parents=True, exist_ok=True)

AUTO_BACKUP_EVERY_HOURS = int(os.getenv("AUTO_BACKUP_EVERY_HOURS", "12") or 12)
BACKUP_RETENTION_COUNT = int(os.getenv("BACKUP_RETENTION_COUNT", "30") or 30)


def ensure_meta_table():
    conn = db()
    try:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS app_meta (
                key TEXT PRIMARY KEY,
                value TEXT
            );
            """
        )
        conn.commit()
    finally:
        conn.close()


def meta_get(key: str, default: str = "") -> str:
    conn = db()
    try:
        row = conn.execute("SELECT value FROM app_meta WHERE key=? LIMIT 1", (key,)).fetchone()
        return str(row["value"]) if row and row["value"] is not None else default
    finally:
        conn.close()


def meta_set(key: str, value: str):
    conn = db()
    try:
        conn.execute(
            """
            INSERT INTO app_meta(key, value)
            VALUES(?, ?)
            ON CONFLICT(key) DO UPDATE SET value=excluded.value
            """,
            (key, str(value)),
        )
        conn.commit()
    finally:
        conn.close()


def list_backups() -> list[Path]:
    files = sorted(BACKUPS_DIR.glob("league_backup_*.db"), reverse=True)
    return files


def _prune_old_backups():
    files = list_backups()
    if BACKUP_RETENTION_COUNT <= 0:
        return
    for p in files[BACKUP_RETENTION_COUNT:]:
        try:
            p.unlink(missing_ok=True)
        except Exception:
            pass


def create_backup_now(label: str = "") -> Path:
    """
    Creates a consistent SQLite backup using the online backup API.
    Returns the backup file path.
    """
    ensure_meta_table()  # ✅ required so meta_set works even on a fresh DB

    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    safe_label = "".join(ch for ch in (label or "").strip() if ch.isalnum() or ch in ("-", "_"))[:24]
    suffix = f"_{safe_label}" if safe_label else ""
    backup_path = BACKUPS_DIR / f"league_backup_{ts}{suffix}.db"

    # Source: live DB (WAL mode)
    src = db()
    try:
        # ✅ checkpoint WAL so backup includes latest data
        src.execute("PRAGMA wal_checkpoint(FULL);")

        # Destination: backup file
        dst = sqlite3.connect(str(backup_path))
        try:
            # ✅ make the backup a normal single-file DB (no WAL)
            dst.execute("PRAGMA journal_mode=DELETE;")

            src.backup(dst)  # consistent snapshot
            dst.commit()
        finally:
            dst.close()
    finally:
        src.close()

    meta_set("last_backup_at", now_iso())
    _prune_old_backups()
    return backup_path


def maybe_auto_backup():
    """
    Auto backup every N hours (default 12). Safe to call every rerun.
    """
    ensure_meta_table()

    last = meta_get("last_auto_backup_at", "")
    do_backup = False

    if not last:
        do_backup = True
    else:
        try:
            last_dt = dtparser.parse(last)
            age = datetime.now(timezone.utc) - last_dt
            do_backup = age.total_seconds() >= (AUTO_BACKUP_EVERY_HOURS * 3600)
        except Exception:
            do_backup = True

    if do_backup:
        p = create_backup_now(label="auto")
        meta_set("last_auto_backup_at", now_iso())
        return p

    return None


def restore_from_backup_file(backup_file_path: Path):
    """
    Restores the CURRENT DB from a backup DB file using SQLite backup API.
    This avoids file-replace/WAL problems and is safe on Render.
    """
    ensure_meta_table()  # ✅ required so meta_set works even on a fresh DB

    # Open src backup DB
    src = sqlite3.connect(str(backup_file_path))
    try:
        # Open dest live DB (direct, not via db() to keep PRAGMAs simple here)
        dst = sqlite3.connect(DB_PATH, check_same_thread=False)
        try:
            # ✅ turn off FK checks during restore copy
            dst.execute("PRAGMA foreign_keys = OFF;")

            # ✅ copy backup -> live
            src.backup(dst)
            dst.commit()

            # ✅ restore runtime PRAGMAs your app expects
            dst.execute("PRAGMA journal_mode = WAL;")
            dst.execute("PRAGMA foreign_keys = ON;")
            dst.execute("PRAGMA wal_checkpoint(FULL);")
            dst.commit()
        finally:
            dst.close()
    finally:
        src.close()

    meta_set("last_restore_at", now_iso())

    # ✅ OPTIONAL but recommended: clear Streamlit caches so UI reflects restored DB instantly
    try:
        st.cache_data.clear()
    except Exception:
        pass


def ensure_referees_phone_column():
    """
    Safe migration: adds referees.phone if it doesn't exist.
    Works on existing DBs without data loss.
    """
    conn = db()
    try:
        cols = conn.execute("PRAGMA table_info(referees);").fetchall()
        col_names = {c["name"] for c in cols}
        if "phone" not in col_names:
            conn.execute("ALTER TABLE referees ADD COLUMN phone TEXT;")
            conn.commit()
    finally:
        conn.close()

def ensure_ladder_tables():
    """
    Safe migrations for ladder system:
    - teams: team name + division + opening_balance
    - game_results: one row per game with admin-entered scoring inputs
    """
    conn = db()
    try:
        cur = conn.cursor()

        # Teams
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS teams (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE,
                division TEXT NOT NULL,
                opening_balance INTEGER NOT NULL DEFAULT 0
            );
            """
        )

        cols = conn.execute("PRAGMA table_info(teams);").fetchall()
        col_names = {c["name"] for c in cols}
        if "opening_balance" not in col_names:
            conn.execute("ALTER TABLE teams ADD COLUMN opening_balance INTEGER NOT NULL DEFAULT 0;")

        # Game results (new installs include default flags)
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS game_results (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                game_id INTEGER NOT NULL UNIQUE,

                home_score INTEGER NOT NULL DEFAULT 0,
                away_score INTEGER NOT NULL DEFAULT 0,

                home_female_tries INTEGER NOT NULL DEFAULT 0,
                away_female_tries INTEGER NOT NULL DEFAULT 0,

                home_conduct INTEGER NOT NULL DEFAULT 0,   -- 0..10
                away_conduct INTEGER NOT NULL DEFAULT 0,   -- 0..10

                home_unstripped INTEGER NOT NULL DEFAULT 0,
                away_unstripped INTEGER NOT NULL DEFAULT 0,

                home_defaulted INTEGER NOT NULL DEFAULT 0,  -- 0/1
                away_defaulted INTEGER NOT NULL DEFAULT 0,  -- 0/1

                updated_at TEXT NOT NULL,

                FOREIGN KEY(game_id) REFERENCES games(id) ON DELETE CASCADE
            );
            """
        )

        # Existing DBs: add the columns if missing
        cols = conn.execute("PRAGMA table_info(game_results);").fetchall()
        gr_cols = {c["name"] for c in cols}

        if "home_defaulted" not in gr_cols:
            conn.execute("ALTER TABLE game_results ADD COLUMN home_defaulted INTEGER NOT NULL DEFAULT 0;")
        if "away_defaulted" not in gr_cols:
            conn.execute("ALTER TABLE game_results ADD COLUMN away_defaulted INTEGER NOT NULL DEFAULT 0;")

        conn.commit()
    finally:
        conn.close()


def init_db():
    conn = db()
    cur = conn.cursor()

    # --- Core tables first (must exist before migrations) ---
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS referees (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            email TEXT NOT NULL UNIQUE,
            active INTEGER NOT NULL DEFAULT 1
        );
        """
    )

    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS games (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            game_key TEXT NOT NULL UNIQUE,
            home_team TEXT NOT NULL,
            away_team TEXT NOT NULL,
            field_name TEXT NOT NULL,
            start_dt TEXT NOT NULL
        );
        """
    )

    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS assignments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            game_id INTEGER NOT NULL,
            slot_no INTEGER NOT NULL,
            referee_id INTEGER,
            status TEXT NOT NULL DEFAULT 'EMPTY',
            updated_at TEXT NOT NULL,
            UNIQUE(game_id, slot_no),
            FOREIGN KEY(game_id) REFERENCES games(id),
            FOREIGN KEY(referee_id) REFERENCES referees(id)
        );
        """
    )

    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS offers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            assignment_id INTEGER NOT NULL,
            token TEXT NOT NULL UNIQUE,
            created_at TEXT NOT NULL,
            responded_at TEXT,
            response TEXT,
            FOREIGN KEY(assignment_id) REFERENCES assignments(id)
        );
        """
    )

    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS blackouts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            referee_id INTEGER NOT NULL,
            blackout_date TEXT NOT NULL,
            UNIQUE(referee_id, blackout_date),
            FOREIGN KEY(referee_id) REFERENCES referees(id)
        );
        """
    )

    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS admins (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            email TEXT NOT NULL UNIQUE,
            active INTEGER NOT NULL DEFAULT 1,
            created_at TEXT NOT NULL
        );
        """
    )

    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS admin_tokens (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            email TEXT NOT NULL,
            token TEXT NOT NULL UNIQUE,
            created_at TEXT NOT NULL,
            expires_at TEXT NOT NULL,
            used_at TEXT
        );
        """
    )

    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS admin_sessions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            email TEXT NOT NULL,
            token TEXT NOT NULL UNIQUE,
            created_at TEXT NOT NULL,
            expires_at TEXT NOT NULL,
            revoked_at TEXT
        );
        """
    )

    conn.commit()
    conn.close()

    # --- Now safe migrations / add-on tables ---
    ensure_referees_phone_column()
    ensure_ladder_tables()


# ============================================================
# Email (SMTP)
# ============================================================
def smtp_settings():
    """
    Required:
      SMTP_HOST, SMTP_PORT, SMTP_USER, SMTP_PASSWORD,
      SMTP_FROM_EMAIL, SMTP_FROM_NAME, APP_BASE_URL
    """
    secrets_dict = {}

    secrets_paths = [
        "/opt/render/.streamlit/secrets.toml",
        "/opt/render/project/src/.streamlit/secrets.toml",
        str(BASE_DIR / ".streamlit" / "secrets.toml"),
    ]
    if any(os.path.exists(p) for p in secrets_paths):
        try:
            secrets_dict = dict(st.secrets)
        except Exception:
            secrets_dict = {}

    def get(key: str, default: str = "") -> str:
        return os.environ.get(key, str(secrets_dict.get(key, default)))

    return {
        "host": get("SMTP_HOST", ""),
        "port": int(get("SMTP_PORT", "587") or 587),
        "user": get("SMTP_USER", ""),
        "password": get("SMTP_PASSWORD", ""),
        "from_email": get("SMTP_FROM_EMAIL", ""),
        "from_name": get("SMTP_FROM_NAME", "Referee Allocator"),
        "app_base_url": get("APP_BASE_URL", "").rstrip("/"),
    }


def send_html_email(
    to_email: str,
    to_name: str,
    subject: str,
    html_body: str,
    text_body: str | None = None,
):
    cfg = smtp_settings()
    if not (
        cfg["host"]
        and cfg["user"]
        and cfg["password"]
        and cfg["from_email"]
        and cfg["app_base_url"]
    ):
        raise RuntimeError(
            "Email not configured. Set SMTP_HOST, SMTP_PORT, SMTP_USER, SMTP_PASSWORD, "
            "SMTP_FROM_EMAIL, SMTP_FROM_NAME, APP_BASE_URL."
        )

    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"] = f'{cfg["from_name"]} <{cfg["from_email"]}>'
    msg["To"] = f"{to_name} <{to_email}>"

    if not text_body:
        text_body = "You have a notification from Referee Allocator."
    msg.attach(MIMEText(text_body, "plain"))
    msg.attach(MIMEText(html_body, "html"))

    with smtplib.SMTP(cfg["host"], cfg["port"]) as server:
        server.starttls()
        server.login(cfg["user"], cfg["password"])
        server.sendmail(cfg["from_email"], [to_email], msg.as_string())


# ============================================================
# Admin auth
# ============================================================
def admin_count() -> int:
    conn = db()
    row = conn.execute("SELECT COUNT(*) AS c FROM admins").fetchone()
    conn.close()
    return int(row["c"])


def add_admin(email: str):
    email = email.strip().lower()
    conn = db()
    conn.execute(
        "INSERT OR IGNORE INTO admins(email, active, created_at) VALUES (?, 1, ?)",
        (email, now_iso()),
    )
    conn.commit()
    conn.close()


def is_admin_email_allowed(email: str) -> bool:
    conn = db()
    row = conn.execute(
        "SELECT 1 FROM admins WHERE email=? AND active=1 LIMIT 1",
        (email.strip().lower(),),
    ).fetchone()
    conn.close()
    return bool(row)


def set_admin_active(email: str, active: bool):
    conn = db()
    conn.execute(
        "UPDATE admins SET active=? WHERE email=?",
        (1 if active else 0, email.strip().lower()),
    )
    conn.commit()
    conn.close()


def list_admins():
    conn = db()
    rows = conn.execute(
        "SELECT email, active, created_at FROM admins ORDER BY email ASC"
    ).fetchall()
    conn.close()
    return rows


def create_admin_login_token(email: str, minutes_valid: int = 15) -> str:
    token = secrets.token_urlsafe(32)
    created = datetime.now(timezone.utc)
    expires = created + timedelta(minutes=minutes_valid)

    conn = db()
    conn.execute(
        """
        INSERT INTO admin_tokens(email, token, created_at, expires_at)
        VALUES (?, ?, ?, ?)
        """,
        (
            email.strip().lower(),
            token,
            created.isoformat(timespec="seconds"),
            expires.isoformat(timespec="seconds"),
        ),
    )
    conn.commit()
    conn.close()
    return token


def consume_admin_token(token: str) -> tuple[bool, str]:
    conn = db()
    row = conn.execute(
        """
        SELECT id, email, expires_at, used_at
        FROM admin_tokens
        WHERE token=?
        """,
        (token,),
    ).fetchone()

    if not row:
        conn.close()
        return False, "Invalid or unknown login link."

    if row["used_at"] is not None:
        conn.close()
        return False, "This login link has already been used."

    expires_at = dtparser.parse(row["expires_at"])
    if datetime.now(timezone.utc) > expires_at:
        conn.close()
        return False, "This login link has expired. Please request a new one."

    email = row["email"].strip().lower()
    if not is_admin_email_allowed(email):
        conn.close()
        return False, "This email is not an authorised administrator."

    conn.execute("UPDATE admin_tokens SET used_at=? WHERE id=?", (now_iso(), row["id"]))
    conn.commit()
    conn.close()
    return True, email


def create_admin_session(email: str, days_valid: int = 14) -> str:
    token = secrets.token_urlsafe(32)
    created = datetime.now(timezone.utc)
    expires = created + timedelta(days=days_valid)

    conn = db()
    conn.execute(
        """
        INSERT INTO admin_sessions(email, token, created_at, expires_at, revoked_at)
        VALUES (?, ?, ?, ?, NULL)
        """,
        (
            email.strip().lower(),
            token,
            created.isoformat(timespec="seconds"),
            expires.isoformat(timespec="seconds"),
        ),
    )
    conn.commit()
    conn.close()
    return token


def consume_admin_session(token: str) -> tuple[bool, str]:
    conn = db()
    row = conn.execute(
        """
        SELECT email, expires_at, revoked_at
        FROM admin_sessions
        WHERE token=?
        LIMIT 1
        """,
        (token,),
    ).fetchone()

    if not row:
        conn.close()
        return False, "Invalid session."

    if row["revoked_at"] is not None:
        conn.close()
        return False, "Session revoked."

    if datetime.now(timezone.utc) > dtparser.parse(row["expires_at"]):
        conn.close()
        return False, "Session expired."

    email = row["email"].strip().lower()
    if not is_admin_email_allowed(email):
        conn.close()
        return False, "Not authorised."

    conn.close()
    return True, email


def maybe_restore_admin_from_session_param():
    qp = st.query_params
    token = qp.get("session")
    if token and not st.session_state.get("admin_email"):
        ok, value = consume_admin_session(token)
        if ok:
            st.session_state["admin_email"] = value
        else:
            st.query_params.pop("session", None)
            st.rerun()


def send_admin_login_email(email: str) -> str:
    email = email.strip().lower()
    cfg = smtp_settings()
    base = cfg.get("app_base_url", "").rstrip("/")
    token = create_admin_login_token(email)
    login_url = f"{base}/?admin_login=1&token={token}"

    subject = "Admin login link"
    text = (
        "Use this link to sign in as an administrator (expires in 15 minutes):\n"
        f"{login_url}\n"
    )
    html = f"""
    <div style="font-family: Arial, sans-serif; line-height: 1.4;">
      <p>Hi,</p>
      <p>Use the button below to sign in as an administrator.
         This link expires in <b>15 minutes</b>.</p>
      <p>
        <a href="{login_url}" style="display:inline-block;padding:10px 14px;background:#1565c0;color:#fff;text-decoration:none;border-radius:6px;">
          Sign in
        </a>
      </p>
      <p>If you didn't request this, you can ignore this email.</p>
    </div>
    """
    send_html_email(email, email, subject, html, text_body=text)
    return login_url


def handle_admin_login_via_query_params():
    qp = st.query_params
    if qp.get("admin_login") == "1" and qp.get("token"):
        token = qp.get("token")
        ok, value = consume_admin_token(token)
        if ok:
            st.session_state["admin_email"] = value
            session_token = create_admin_session(value)

            st.query_params.pop("admin_login", None)
            st.query_params.pop("token", None)
            st.query_params["session"] = session_token
            st.rerun()
        else:
            st.title("Admin Login")
            st.error(value)
            st.info("Please go back and request a new login link.")
            st.stop()


def admin_logout_button():
    if st.session_state.get("admin_email"):
        c1, c2 = st.columns([3, 1])
        with c1:
            st.caption(f"Logged in as: {st.session_state['admin_email']}")
        with c2:
            if st.button("Log out"):
                st.session_state.pop("admin_email", None)
                st.query_params.pop("session", None)
                st.rerun()


def create_admin_session_with_expires_at(email: str, expires_at_iso: str) -> str:
    """
    DEV helper: creates an admin session with a fixed expiry.
    Used for permanent DEV admin URLs (no email).
    """
    token = secrets.token_urlsafe(32)

    conn = db()
    conn.execute(
        """
        INSERT INTO admin_sessions(email, token, created_at, expires_at, revoked_at)
        VALUES (?, ?, ?, ?, NULL)
        """,
        (
            email.strip().lower(),
            token,
            now_iso(),
            expires_at_iso,
        ),
    )
    conn.commit()
    conn.close()

    return token


# ============================================================
# Imports & data helpers
# ============================================================
def import_referees_csv(df: pd.DataFrame):
    cols = {c.lower().strip(): c for c in df.columns}
    if "name" not in cols or "email" not in cols:
        raise ValueError("Referees CSV must contain columns: name, email")

    phone_col = cols.get("phone")  # optional

    conn = db()
    cur = conn.cursor()
    added = 0
    updated = 0

    for _, row in df.iterrows():
        name = str(row[cols["name"]]).strip()
        email = str(row[cols["email"]]).strip().lower()
        phone = ""
        if phone_col:
            phone = str(row[phone_col]).strip()
            if phone.lower() == "nan":
                phone = ""

        if not name or not email or email == "nan":
            continue

        cur.execute("SELECT id, name, COALESCE(phone,'') AS phone FROM referees WHERE email=?", (email,))
        existing = cur.fetchone()

        if existing:
            needs_update = False
            if existing["name"] != name:
                needs_update = True
            if phone_col and (existing["phone"] or "") != phone:
                needs_update = True

            if needs_update:
                cur.execute(
                    "UPDATE referees SET name=?, phone=? WHERE email=?",
                    (name, phone, email),
                )
                updated += 1
        else:
            cur.execute(
                "INSERT INTO referees(name, email, phone, active) VALUES (?, ?, ?, 1)",
                (name, email, phone),
            )
            added += 1

    conn.commit()
    conn.close()
    return added, updated


def replace_referees_csv(df: pd.DataFrame):
    cols = {c.lower().strip(): c for c in df.columns}
    if "name" not in cols or "email" not in cols:
        raise ValueError("Referees CSV must contain columns: name, email")

    phone_col = cols.get("phone")  # optional

    new_refs = []
    for _, row in df.iterrows():
        name = str(row[cols["name"]]).strip()
        email = str(row[cols["email"]]).strip().lower()

        phone = ""
        if phone_col:
            phone = str(row[phone_col]).strip()
            if phone.lower() == "nan":
                phone = ""

        if not name or not email or email == "nan":
            continue

        new_refs.append((name, email, phone))

    if len(new_refs) == 0:
        raise ValueError("Referees CSV has no valid rows. Aborting replace import (nothing deleted).")

    conn = db()
    cur = conn.cursor()
    try:
        cur.execute("BEGIN")
        cur.execute("DELETE FROM offers")
        cur.execute(
            """
            UPDATE assignments
            SET referee_id=NULL, status='EMPTY', updated_at=?
            """,
            (now_iso(),),
        )
        cur.execute("DELETE FROM blackouts")
        cur.execute("DELETE FROM referees")

        try:
            cur.execute("DELETE FROM sqlite_sequence WHERE name IN ('referees','blackouts','offers')")
        except Exception:
            pass

        cur.executemany(
            "INSERT INTO referees(name, email, phone, active) VALUES (?, ?, ?, 1)",
            new_refs,
        )
        conn.commit()
        return len(new_refs)
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def import_games_csv(df: pd.DataFrame):
    cols = {c.lower().strip(): c for c in df.columns}

    key_col = None
    for k in ["game_id", "game_key", "id"]:
        if k in cols:
            key_col = cols[k]
            break

    has_date_time = "date" in cols and "start_time" in cols
    has_start_dt = "start_datetime" in cols

    required = ["home_team", "away_team", "field"]
    missing = [r for r in required if r not in cols]
    if not key_col:
        missing.append("game_id")
    if not (has_date_time or has_start_dt):
        missing.append("date + start_time (or start_datetime)")

    if missing:
        raise ValueError(f"Games CSV missing columns: {', '.join(missing)}")

    conn = db()
    cur = conn.cursor()

    inserted, updated = 0, 0
    for _, row in df.iterrows():
        game_key = str(row[key_col]).strip()
        home = str(row[cols["home_team"]]).strip()
        away = str(row[cols["away_team"]]).strip()
        field = str(row[cols["field"]]).strip()

        if not (game_key and home and away and field) or game_key == "nan":
            continue

        try:
            if has_date_time:
                d_raw = str(row[cols["date"]]).strip()
                t_raw = str(row[cols["start_time"]]).strip()
                start_dt = dtparser.parse(f"{d_raw} {t_raw}")
            else:
                start_raw = str(row[cols["start_datetime"]]).strip()
                d = parse_csv_date_strict(d_raw)

                # parse time strictly-ish (handles "18:00" or "6:00 PM")
                t = dtparser.parse(t_raw).time()

                start_dt = datetime.combine(d, t)

        except Exception as e:
            raise ValueError(f"Could not parse date/time for game_id={game_key}: {e}")

        start_iso = start_dt.isoformat(timespec="minutes")

        cur.execute("SELECT id FROM games WHERE game_key=?", (game_key,))
        g = cur.fetchone()

        if g:
            cur.execute(
                """
                UPDATE games
                SET home_team=?, away_team=?, field_name=?, start_dt=?
                WHERE game_key=?
                """,
                (home, away, field, start_iso, game_key),
            )
            updated += 1
            game_id = g["id"]
        else:
            cur.execute(
                """
                INSERT INTO games(game_key, home_team, away_team, field_name, start_dt)
                VALUES (?, ?, ?, ?, ?)
                """,
                (game_key, home, away, field, start_iso),
            )
            inserted += 1
            game_id = cur.lastrowid

        for slot in (1, 2):
            cur.execute(
                "SELECT 1 FROM assignments WHERE game_id=? AND slot_no=?",
                (game_id, slot),
            )
            if not cur.fetchone():
                cur.execute(
                    """
                    INSERT INTO assignments(game_id, slot_no, referee_id, status, updated_at)
                    VALUES (?, ?, NULL, 'EMPTY', ?)
                    """,
                    (game_id, slot, now_iso()),
                )

    conn.commit()
    conn.close()
    return inserted, updated

def replace_games_csv(df: pd.DataFrame):
    """
    REPLACE mode: deletes ALL games (and dependent data) then imports the CSV as the new draw.

    What gets reset:
      - offers (must be first; depends on assignments)
      - assignments
      - game_results (ladder scoring inputs)
      - games

    What is preserved:
      - referees
      - blackouts
      - teams (divisions/opening balances)
      - admins/admin sessions/tokens

    Returns: (imported_games_count)
    """
    cols = {c.lower().strip(): c for c in df.columns}

    key_col = None
    for k in ["game_id", "game_key", "id"]:
        if k in cols:
            key_col = cols[k]
            break

    has_date_time = "date" in cols and "start_time" in cols
    has_start_dt = "start_datetime" in cols

    required = ["home_team", "away_team", "field"]
    missing = [r for r in required if r not in cols]
    if not key_col:
        missing.append("game_id")
    if not (has_date_time or has_start_dt):
        missing.append("date + start_time (or start_datetime)")

    if missing:
        raise ValueError(f"Games CSV missing columns: {', '.join(missing)}")

    # Build a clean list of new games first (so we never delete if CSV is junk)
    new_games = []
    seen_keys = set()

    for _, row in df.iterrows():
        game_key = str(row[key_col]).strip()
        home = str(row[cols["home_team"]]).strip()
        away = str(row[cols["away_team"]]).strip()
        field = str(row[cols["field"]]).strip()

        if not (game_key and home and away and field) or game_key.lower() == "nan":
            continue

        if game_key in seen_keys:
            raise ValueError(f"Duplicate game_id/game_key found in CSV: {game_key}")
        seen_keys.add(game_key)

        try:
            if has_date_time:
                d_raw = str(row[cols["date"]]).strip()
                t_raw = str(row[cols["start_time"]]).strip()
                
                d = parse_csv_date_strict(d_raw)
                # parse time strictly-ish (handles "18:00" or "6:00 PM")
                t = dtparser.parse(t_raw).time()
                start_dt = datetime.combine(d, t)
            else:
                start_raw = str(row[cols["start_datetime"]]).strip()
                start_dt = dtparser.parse(start_raw)
        except Exception as e:
            raise ValueError(f"Could not parse date/time for game_id={game_key}: {e}")

        start_iso = start_dt.isoformat(timespec="minutes")
        new_games.append((game_key, home, away, field, start_iso))

    if len(new_games) == 0:
        raise ValueError("Games CSV has no valid rows. Aborting replace import (nothing deleted).")

    conn = db()
    cur = conn.cursor()
    try:
        cur.execute("BEGIN")

        # Delete in dependency order
        cur.execute("DELETE FROM offers")
        cur.execute("DELETE FROM assignments")
        cur.execute("DELETE FROM game_results")
        cur.execute("DELETE FROM games")

        # Optional: reset autoincrement counters (safe if table exists)
        try:
            cur.execute(
                "DELETE FROM sqlite_sequence WHERE name IN ('games','assignments','offers','game_results')"
            )
        except Exception:
            pass

        # Insert new games + their 2 empty assignments
        imported = 0
        for game_key, home, away, field, start_iso in new_games:
            cur.execute(
                """
                INSERT INTO games(game_key, home_team, away_team, field_name, start_dt)
                VALUES (?, ?, ?, ?, ?)
                """,
                (game_key, home, away, field, start_iso),
            )
            game_id = cur.lastrowid

            for slot in (1, 2):
                cur.execute(
                    """
                    INSERT INTO assignments(game_id, slot_no, referee_id, status, updated_at)
                    VALUES (?, ?, NULL, 'EMPTY', ?)
                    """,
                    (game_id, slot, now_iso()),
                )

            imported += 1

        conn.commit()
        return imported

    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def import_blackouts_csv(df: pd.DataFrame):
    cols = {c.lower().strip(): c for c in df.columns}
    if "email" not in cols or "blackout_date" not in cols:
        raise ValueError("Blackouts CSV must contain columns: email, blackout_date")

    conn = db()
    cur = conn.cursor()

    added, skipped = 0, 0
    for _, row in df.iterrows():
        email = str(row[cols["email"]]).strip().lower()
        d_raw = str(row[cols["blackout_date"]]).strip()
        if not email or email == "nan" or not d_raw or d_raw == "nan":
            continue

        cur.execute("SELECT id FROM referees WHERE email=?", (email,))
        r = cur.fetchone()
        if not r:
            skipped += 1
            continue

        try:
            d = dtparser.parse(d_raw).date()
        except Exception:
            raise ValueError(f"Could not parse blackout_date '{d_raw}' for {email}")

        try:
            cur.execute(
                "INSERT INTO blackouts(referee_id, blackout_date) VALUES (?, ?)",
                (r["id"], d.isoformat()),
            )
            added += 1
        except sqlite3.IntegrityError:
            pass

    conn.commit()
    conn.close()
    return added, skipped


def list_all_blackouts_with_ref_names() -> pd.DataFrame:
    """
    Returns all blackout dates with referee names, sorted by date then name.
    Inserts a blank row between dates for readability in the sidebar table.
    """
    conn = db()
    rows = conn.execute(
        """
        SELECT
            b.blackout_date AS blackout_date,
            TRIM(r.name) AS referee
        FROM blackouts b
        JOIN referees r ON r.id = b.referee_id
        ORDER BY b.blackout_date ASC, r.name ASC
        """
    ).fetchall()
    conn.close()

    if not rows:
        return pd.DataFrame(columns=["Date", "Referee"])

    out = []
    last_date = None
    for r in rows:
        d = (r["blackout_date"] or "").strip()
        nm = (r["referee"] or "").strip()

        if last_date is not None and d != last_date:
            out.append({"Date": "", "Referee": ""})  # spacer row

        out.append({"Date": d, "Referee": nm})
        last_date = d

    return pd.DataFrame(out)


def get_games():
    conn = db()
    rows = conn.execute(
        """
        SELECT id, game_key, home_team, away_team, field_name, start_dt
        FROM games
        ORDER BY start_dt ASC
        """
    ).fetchall()
    conn.close()
    return rows


def get_referees():
    conn = db()
    rows = conn.execute(
        """
        SELECT id, name, email, COALESCE(phone,'') AS phone
        FROM referees
        WHERE active=1
        ORDER BY name ASC
        """
    ).fetchall()
    conn.close()
    return rows


def get_assignments_for_game(game_id: int):
    conn = db()
    rows = conn.execute(
        """
        SELECT
            a.id,
            a.slot_no,
            a.referee_id,
            a.status,
            a.updated_at,
            r.name AS ref_name,
            r.email AS ref_email
        FROM assignments a
        LEFT JOIN referees r ON r.id = a.referee_id
        WHERE a.game_id=?
        ORDER BY a.slot_no ASC
        """,
        (game_id,),
    ).fetchall()
    conn.close()
    return rows

def get_assignment_live(assignment_id: int):
    """
    Re-fetch the latest assignment row (and linked referee details) from the DB.

    Used inside Action handlers so we don't rely on stale 'a' from the UI loop.
    Returns sqlite3.Row or None.
    """
    conn = db()
    row = conn.execute(
        """
        SELECT
            a.id,
            a.game_id,
            a.slot_no,
            a.referee_id,
            a.status,
            a.updated_at,
            r.name  AS ref_name,
            r.email AS ref_email,
            COALESCE(r.phone,'') AS ref_phone
        FROM assignments a
        LEFT JOIN referees r ON r.id = a.referee_id
        WHERE a.id = ?
        LIMIT 1
        """,
        (int(assignment_id),),
    ).fetchone()
    conn.close()
    return row

# ============================================================
# Assignment helpers
# ============================================================

def set_assignment_status(assignment_id: int, status: str):
    status = (status or "EMPTY").strip().upper()
    conn = db()
    conn.execute(
        """
        UPDATE assignments
        SET status=?, updated_at=?
        WHERE id=?
        """,
        (status, now_iso(), int(assignment_id)),
    )
    conn.commit()
    conn.close()


def set_assignment_ref(assignment_id: int, referee_id: int):
    """
    Set the referee for a slot.
    If a referee is set and the slot was EMPTY, we move it to NOT_OFFERED
    (your UI treats any non-empty referee as 'NOT OFFERED YET' unless OFFERED/DECLINED/ACCEPTED/ASSIGNED).
    """
    conn = db()

    # Keep status consistent with your UI badges
    cur = conn.execute(
        "SELECT referee_id, status FROM assignments WHERE id=? LIMIT 1",
        (int(assignment_id),),
    ).fetchone()

    if not cur:
        conn.close()
        return

    cur_status = (cur["status"] or "EMPTY").strip().upper()
    new_status = cur_status

    if cur_status == "EMPTY":
        new_status = "NOT_OFFERED"

    conn.execute(
        """
        UPDATE assignments
        SET referee_id=?, status=?, updated_at=?
        WHERE id=?
        """,
        (int(referee_id), new_status, now_iso(), int(assignment_id)),
    )
    conn.commit()
    conn.close()


def _delete_offers_for_assignment(conn: sqlite3.Connection, assignment_id: int):
    """
    Internal helper: ensure no stale offer links remain for this assignment.
    """
    conn.execute("DELETE FROM offers WHERE assignment_id=?", (int(assignment_id),))


def clear_assignment(assignment_id: int):
    """
    Clears the slot back to EMPTY and removes any offers for that assignment.
    This is what your RESET/DELETE action needs.
    """
    conn = db()
    try:
        _delete_offers_for_assignment(conn, assignment_id)
        conn.execute(
            """
            UPDATE assignments
            SET referee_id=NULL, status='EMPTY', updated_at=?
            WHERE id=?
            """,
            (now_iso(), int(assignment_id)),
        )
        conn.commit()
    finally:
        conn.close()

        
# ============================================================
# Ladder / scoring helpers
# ============================================================

LADDER_WIN_PTS = 3
LADDER_DRAW_PTS = 2
LADDER_LOSS_PTS = 0

DIVISIONS = [
    "Division 1",
    "Division 2",
    "Division 3",
    "Golden Oldies",
    "Other",
]


def upsert_team(name: str, division: str, opening_balance: int = 0):
    name = (name or "").strip()
    division = (division or "").strip()
    if not name or not division:
        return

    conn = db()
    try:
        conn.execute(
            """
            INSERT INTO teams(name, division, opening_balance)
            VALUES (?, ?, ?)
            ON CONFLICT(name) DO UPDATE SET
                division=excluded.division,
                opening_balance=excluded.opening_balance
            """,
            (name, division, int(opening_balance or 0)),
        )
        conn.commit()
    finally:
        conn.close()


def list_teams() -> list[sqlite3.Row]:
    conn = db()
    rows = conn.execute(
        """
        SELECT
            id,
            name,
            division,
            COALESCE(opening_balance,0) AS opening_balance
        FROM teams
        ORDER BY division ASC, name ASC
        """
    ).fetchall()
    conn.close()
    return rows


def get_team_division(name: str) -> str:
    conn = db()
    row = conn.execute(
        "SELECT division FROM teams WHERE name=? LIMIT 1",
        ((name or "").strip(),),
    ).fetchone()
    conn.close()
    return (row["division"] if row else "").strip()


def get_game_result(game_id: int) -> sqlite3.Row | None:
    conn = db()
    row = conn.execute(
        """
        SELECT
            game_id,
            home_score, away_score,
            home_female_tries, away_female_tries,
            home_conduct, away_conduct,
            home_unstripped, away_unstripped,
            COALESCE(home_defaulted,0) AS home_defaulted,
            COALESCE(away_defaulted,0) AS away_defaulted,
            updated_at
        FROM game_results
        WHERE game_id=?
        LIMIT 1
        """,
        (game_id,),
    ).fetchone()
    conn.close()
    return row


def upsert_game_result(
    *,
    game_id: int,
    home_score: int,
    away_score: int,
    home_female_tries: int,
    away_female_tries: int,
    home_conduct: int,
    away_conduct: int,
    home_unstripped: int,
    away_unstripped: int,
    home_defaulted: int = 0,
    away_defaulted: int = 0,
):
    conn = db()
    try:
        conn.execute(
            """
            INSERT INTO game_results(
                game_id,
                home_score, away_score,
                home_female_tries, away_female_tries,
                home_conduct, away_conduct,
                home_unstripped, away_unstripped,
                home_defaulted, away_defaulted,
                updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(game_id) DO UPDATE SET
                home_score=excluded.home_score,
                away_score=excluded.away_score,
                home_female_tries=excluded.home_female_tries,
                away_female_tries=excluded.away_female_tries,
                home_conduct=excluded.home_conduct,
                away_conduct=excluded.away_conduct,
                home_unstripped=excluded.home_unstripped,
                away_unstripped=excluded.away_unstripped,
                home_defaulted=excluded.home_defaulted,
                away_defaulted=excluded.away_defaulted,
                updated_at=excluded.updated_at
            """,
            (
                int(game_id),
                int(home_score), int(away_score),
                int(home_female_tries), int(away_female_tries),
                int(home_conduct), int(away_conduct),
                int(home_unstripped), int(away_unstripped),
                1 if int(home_defaulted or 0) else 0,
                1 if int(away_defaulted or 0) else 0,
                now_iso(),
            ),
        )
        conn.commit()
    finally:
        conn.close()


def ladder_audit_df_as_at(as_at_date: date) -> pd.DataFrame:
    """
    Audit table: one row per TEAM per GAME
    from season start → as_at_date.

    IMPORTANT:
    - If a game has NO saved result row in game_results, it contributes NOTHING (0 points, not played).
      This prevents unsaved games being treated as 0-0 draws worth 2 points.
    """
    season_start = get_season_start_date()
    if not season_start:
        return pd.DataFrame()

    start_min = datetime.combine(season_start, datetime.min.time()).isoformat(timespec="seconds")
    end_max = datetime.combine(as_at_date + timedelta(days=1), datetime.min.time()).isoformat(timespec="seconds")

    conn = db()
    rows = conn.execute(
        """
        WITH base AS (
            SELECT
                g.id AS game_id,
                g.start_dt,
                g.field_name,
                g.home_team,
                g.away_team,

                COALESCE(t1.division,'—') AS home_division,
                COALESCE(t2.division,'—') AS away_division,
                COALESCE(t1.opening_balance,0) AS home_opening,
                COALESCE(t2.opening_balance,0) AS away_opening,

                -- if gr.game_id is NULL, there is NO saved result
                gr.game_id AS has_result,

                gr.home_score, gr.away_score,
                gr.home_female_tries, gr.away_female_tries,
                gr.home_conduct, gr.away_conduct,
                gr.home_unstripped, gr.away_unstripped,
                COALESCE(gr.home_defaulted,0) AS home_defaulted,
                COALESCE(gr.away_defaulted,0) AS away_defaulted
            FROM games g
            LEFT JOIN teams t1 ON t1.name = g.home_team
            LEFT JOIN teams t2 ON t2.name = g.away_team
            LEFT JOIN game_results gr ON gr.game_id = g.id
            WHERE g.start_dt >= ? AND g.start_dt < ?
        ),
        teamsplit AS (
            SELECT
                game_id, start_dt,
                home_team AS team,
                away_team AS opponent,
                home_division AS division,
                home_opening AS opening,

                has_result,

                home_score AS pf,
                away_score AS pa,
                home_female_tries AS female_tries,
                home_conduct AS conduct,
                home_unstripped AS unstripped,
                home_defaulted AS defaulted
            FROM base

            UNION ALL

            SELECT
                game_id, start_dt,
                away_team AS team,
                home_team AS opponent,
                away_division AS division,
                away_opening AS opening,

                has_result,

                away_score AS pf,
                home_score AS pa,
                away_female_tries AS female_tries,
                away_conduct AS conduct,
                away_unstripped AS unstripped,
                away_defaulted AS defaulted
            FROM base
        )
        SELECT * FROM teamsplit
        ORDER BY start_dt, team
        """,
        (start_min, end_max),
    ).fetchall()
    conn.close()

    out = []
    for r in rows:
        # ✅ key rule: if no saved result row, ignore this game completely
        if r["has_result"] is None:
            continue

        pf = int(r["pf"] or 0)
        pa = int(r["pa"] or 0)

        bd = compute_points_breakdown_for_game(
            home_score=pf,
            away_score=pa,
            home_female_tries=int(r["female_tries"] or 0),
            away_female_tries=0,  # unused in single-team row
            home_conduct=int(r["conduct"] or 0),
            away_conduct=0,       # unused
            home_unstripped=int(r["unstripped"] or 0),
            away_unstripped=0,    # unused
            home_defaulted=int(r["defaulted"] or 0),
            away_defaulted=0,
        )["HOME"]

        out.append({
            "Date": (r["start_dt"] or "")[:10],
            "Division": r["division"],
            "Team": r["team"],
            "Opponent": r["opponent"],

            "PF": pf,
            "PA": pa,
            "PD": pf - pa,

            "Res": bd.get("Res", ""),

            "Match": int(bd.get("Match", 0)),
            "CloseBP": int(bd.get("CloseBP", 0)),
            "FemBP": int(bd.get("FemBP", 0)),
            "Conduct": int(bd.get("Conduct", 0)),
            "Pen": int(bd.get("Pen", 0)),

            "Points": int(bd.get("Points", 0)),
            "Opening": int(r["opening"] or 0),

            "Defaulted": bool(bd.get("Defaulted", False)),
        })

    return pd.DataFrame(out)



def ladder_table_df_as_at(as_at_date: date, division: str) -> pd.DataFrame:
    # Base list of teams in this division (so they show even if no results saved yet)
    teams = [r for r in list_teams() if (r["division"] or "").strip() == division]
    if not teams:
        return pd.DataFrame()

    base = pd.DataFrame([
        {"Team": r["name"], "Opening": int(r["opening_balance"] or 0)}
        for r in teams
    ])

    df = ladder_audit_df_as_at(as_at_date)
    if df.empty:
        # No saved results yet → ladder is just opening balances
        base["P"] = 0
        base["W"] = 0
        base["D"] = 0
        base["L"] = 0
        base["PF"] = 0
        base["PA"] = 0
        base["PD"] = 0
        base["Points"] = 0
        base["Total"] = base["Opening"]
        return base.sort_values(by=["Total", "Team"], ascending=[False, True]).reset_index(drop=True)

    df = df[df["Division"] == division].copy()
    if df.empty:
        base["P"] = 0
        base["W"] = 0
        base["D"] = 0
        base["L"] = 0
        base["PF"] = 0
        base["PA"] = 0
        base["PD"] = 0
        base["Points"] = 0
        base["Total"] = base["Opening"]
        return base.sort_values(by=["Total", "Team"], ascending=[False, True]).reset_index(drop=True)

    grouped = df.groupby("Team", dropna=False).agg(
        P=("Res", lambda s: s.isin(["W", "D", "L"]).sum()),
        W=("Res", lambda s: (s == "W").sum()),
        D=("Res", lambda s: (s == "D").sum()),
        L=("Res", lambda s: (s == "L").sum()),
        PF=("PF", "sum"),
        PA=("PA", "sum"),
        Points=("Points", "sum"),
    ).reset_index()

    merged = base.merge(grouped, on="Team", how="left").fillna(0)

    # fix numeric dtypes after fillna
    for c in ["P", "W", "D", "L", "PF", "PA", "Points", "Opening"]:
        merged[c] = merged[c].astype(int)

    merged["PD"] = merged["PF"] - merged["PA"]
    merged["Total"] = merged["Opening"] + merged["Points"]

    return merged.sort_values(
        by=["Total", "PD", "PF", "Team"],
        ascending=[False, False, False, True],
    ).reset_index(drop=True)


def compute_points_breakdown_for_game(
    *,
    home_score: int,
    away_score: int,
    home_female_tries: int,
    away_female_tries: int,
    home_conduct: int,
    away_conduct: int,
    home_unstripped: int,
    away_unstripped: int,
    home_defaulted: int,
    away_defaulted: int,
) -> dict:
    """
    Computes a full points breakdown for HOME and AWAY teams.
    Explicitly marks Defaulted=True/False for UI use.
    """

    def side_breakdown(
        *,
        pf: int,
        pa: int,
        female_tries: int,
        conduct: int,
        unstripped: int,
        defaulted: int,
        is_home: bool,
    ) -> dict:

        is_defaulted = bool(defaulted)

        # Match result + match points
        if is_defaulted:
            res = "DEFAULT"
            match_pts = 0
        elif pf > pa:
            res = "W"
            match_pts = LADDER_WIN_PTS
        elif pf == pa:
            res = "D"
            match_pts = LADDER_DRAW_PTS
        else:
            res = "L"
            match_pts = LADDER_LOSS_PTS

        # Close loss BP
        close_bp = 0
        if not is_defaulted and pf < pa and (pa - pf) in (1, 2):
            close_bp = 1

        # Female BP
        fem_bp = 1 if (not is_defaulted and female_tries >= 4) else 0

        # Unstripped penalty
        pen = -2 if (not is_defaulted and unstripped >= 3) else 0

        total = match_pts + close_bp + fem_bp + conduct + pen

        return {
            "PF": pf,
            "PA": pa,
            "Res": res,
            "Match": match_pts,
            "CloseBP": close_bp,
            "FemTries": female_tries,
            "FemBP": fem_bp,
            "Conduct": conduct,
            "Unstrip": unstripped,
            "Pen": pen,
            "Points": total,
            "Defaulted": is_defaulted,   # ✅ THE IMPORTANT BIT
        }

    home = side_breakdown(
        pf=home_score,
        pa=away_score,
        female_tries=home_female_tries,
        conduct=home_conduct,
        unstripped=home_unstripped,
        defaulted=home_defaulted,
        is_home=True,
    )

    away = side_breakdown(
        pf=away_score,
        pa=home_score,
        female_tries=away_female_tries,
        conduct=away_conduct,
        unstripped=away_unstripped,
        defaulted=away_defaulted,
        is_home=False,
    )

    return {
        "HOME": home,
        "AWAY": away,
    }

# ============================================================
# Acceptance progress helpers
# ============================================================
def iso_week_window(d: date) -> tuple[date, date]:
    start = d - timedelta(days=d.weekday())  # Monday
    end_excl = start + timedelta(days=7)
    return start, end_excl


def get_acceptance_progress_for_window(start_date: date, end_date_exclusive: date) -> tuple[int, int]:
    start_min = datetime.combine(start_date, datetime.min.time()).isoformat(timespec="seconds")
    start_max = datetime.combine(end_date_exclusive, datetime.min.time()).isoformat(timespec="seconds")

    conn = db()
    row = conn.execute(
        """
        SELECT
            SUM(CASE WHEN UPPER(COALESCE(a.status,'')) IN ('ACCEPTED','ASSIGNED') THEN 1 ELSE 0 END) AS accepted_slots,
            COUNT(a.id) AS total_slots
        FROM games g
        JOIN assignments a ON a.game_id = g.id
        WHERE g.start_dt >= ? AND g.start_dt < ?
        """,
        (start_min, start_max),
    ).fetchone()
    conn.close()

    accepted = int(row["accepted_slots"] or 0)
    total = int(row["total_slots"] or 0)
    return accepted, total


def list_referees_not_accepted_for_window(start_date: date, end_date_exclusive: date) -> list[str]:
    start_min = datetime.combine(start_date, datetime.min.time()).isoformat(timespec="seconds")
    start_max = datetime.combine(end_date_exclusive, datetime.min.time()).isoformat(timespec="seconds")

    conn = db()
    rows = conn.execute(
        """
        SELECT DISTINCT TRIM(r.name) AS name
        FROM games g
        JOIN assignments a ON a.game_id = g.id
        JOIN referees r ON r.id = a.referee_id
        WHERE g.start_dt >= ? AND g.start_dt < ?
          AND a.referee_id IS NOT NULL
          AND UPPER(COALESCE(a.status,'')) NOT IN ('ACCEPTED','ASSIGNED')
          AND TRIM(COALESCE(r.name,'')) <> ''
        ORDER BY name ASC
        """,
        (start_min, start_max),
    ).fetchall()
    conn.close()

    return [row["name"] for row in rows]


def has_any_offers_for_window(start_date: date, end_date_exclusive: date) -> bool:
    """
    True if at least one offer exists for any assignment whose game start_dt falls within the window.
    This is the cleanest signal that an OFFER has been sent (or at least created).
    """
    start_min = datetime.combine(start_date, datetime.min.time()).isoformat(timespec="seconds")
    start_max = datetime.combine(end_date_exclusive, datetime.min.time()).isoformat(timespec="seconds")

    conn = db()
    row = conn.execute(
        """
        SELECT 1
        FROM games g
        JOIN assignments a ON a.game_id = g.id
        JOIN offers o ON o.assignment_id = a.id
        WHERE g.start_dt >= ? AND g.start_dt < ?
        LIMIT 1
        """,
        (start_min, start_max),
    ).fetchone()
    conn.close()
    return bool(row)

def get_referee_workload_all_time() -> pd.DataFrame:
    """
    Returns a dataframe of all active referees and how many slots they have that are
    ACCEPTED/ASSIGNED across ALL games in the database.
    Sorted least -> most.
    """
    conn = db()
    rows = conn.execute(
        """
        SELECT
            r.id AS referee_id,
            TRIM(r.name) AS name,
            TRIM(r.email) AS email,
            TRIM(COALESCE(r.phone,'')) AS phone,
            SUM(
                CASE
                    WHEN UPPER(COALESCE(a.status,'')) IN ('ACCEPTED','ASSIGNED')
                    THEN 1 ELSE 0
                END
            ) AS accepted_slots
        FROM referees r
        LEFT JOIN assignments a
            ON a.referee_id = r.id
        WHERE r.active = 1
        GROUP BY r.id, r.name, r.email, r.phone
        ORDER BY accepted_slots ASC, name ASC
        """
    ).fetchall()
    conn.close()

    df = pd.DataFrame(
        [
            {
                "Referee": (row["name"] or "").strip() or "—",
                "Phone": (row["phone"] or "").strip() or "—",
                "Email": (row["email"] or "").strip() or "—",
                "Accepted": int(row["accepted_slots"] or 0),
            }
            for row in rows
        ]
    )

    if df.empty:
        df = pd.DataFrame(columns=["Referee", "Phone", "Email", "Accepted"])

    return df


# ============================================================
# Printable PDF helpers
# ============================================================
def get_admin_print_rows_for_date(selected_date: date):
    start_min = datetime.combine(selected_date, datetime.min.time()).isoformat(timespec="seconds")
    start_max = datetime.combine(selected_date + timedelta(days=1), datetime.min.time()).isoformat(timespec="seconds")

    conn = db()
    rows = conn.execute(
        """
        SELECT
            g.id AS game_id,
            g.home_team,
            g.away_team,
            g.field_name,
            g.start_dt,
            a.slot_no,
            a.status,
            r.name AS ref_name
        FROM games g
        LEFT JOIN assignments a ON a.game_id = g.id
        LEFT JOIN referees r ON r.id = a.referee_id
        WHERE g.start_dt >= ? AND g.start_dt < ?
        ORDER BY g.start_dt ASC, g.field_name ASC, g.home_team ASC, a.slot_no ASC
        """,
        (start_min, start_max),
    ).fetchall()
    conn.close()

    games_map = {}
    for row in rows:
        gid = int(row["game_id"])
        if gid not in games_map:
            games_map[gid] = {
                "home_team": row["home_team"],
                "away_team": row["away_team"],
                "field_name": row["field_name"],
                "start_dt": row["start_dt"],
                "slots": {
                    1: {"name": "", "status": "EMPTY"},
                    2: {"name": "", "status": "EMPTY"},
                },
            }

        slot_no = row["slot_no"]
        if slot_no in (1, 2):
            nm = (row["ref_name"] or "").strip()
            stt = (row["status"] or "EMPTY").strip().upper()
            games_map[gid]["slots"][int(slot_no)] = {"name": nm, "status": stt}

    out = list(games_map.values())
    out.sort(key=lambda x: x["start_dt"])
    return out


def _format_ref_name(name: str, status: str) -> str:
    name = (name or "").strip()
    status = (status or "EMPTY").strip().upper()
    if not name:
        return "—"
    if status in ("ACCEPTED", "ASSIGNED"):
        return name
    if status in ("OFFERED", "DECLINED"):
        return f"{name} ({status})"
    return name


def build_admin_summary_pdf_bytes(selected_date: date) -> bytes:
    games = get_admin_print_rows_for_date(selected_date)

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=16,
        rightMargin=16,
        topMargin=16,
        bottomMargin=16,
        title=f"Game Summary {selected_date.isoformat()}",
    )

    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph(f"<b>Game Summary</b> — {selected_date.isoformat()}", styles["Title"]))
    story.append(Spacer(1, 4))

    if not games:
        story.append(Paragraph("No games found for this date.", styles["Normal"]))
        doc.build(story)
        return buffer.getvalue()

    grouped = {}
    for g in games:
        dt = dtparser.parse(g["start_dt"])
        key = _time_12h(dt)
        grouped.setdefault(key, []).append((dt, g))

    group_keys = sorted(grouped.keys(), key=lambda k: min(dt for dt, _ in grouped[k]))

    for time_key in group_keys:
        story.append(Paragraph(f"<b>Start time: {time_key}</b>", styles["Heading3"]))
        story.append(Spacer(1, 2))

        data = [["Teams", "Field", "Start", "Referees"]]
        for dt, g in grouped[time_key]:
            teams = f"{g['home_team']} vs {g['away_team']}"
            field = g["field_name"]
            start_str = _time_12h(dt)

            r1 = _format_ref_name(g["slots"][1]["name"], g["slots"][1]["status"])
            r2 = _format_ref_name(g["slots"][2]["name"], g["slots"][2]["status"])
            refs = f"{r1} / {r2}"

            data.append([teams, field, start_str, refs])

        table = Table(
            data,
            colWidths=[360, 110, 75, 210],
            repeatRows=1,
        )
        table.setStyle(
            TableStyle(
                [
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 9),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                    ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                    ("FONTSIZE", (0, 1), (-1, -1), 8),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 3),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                    ("TOPPADDING", (0, 0), (-1, -1), 1),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
                ]
            )
        )

        story.append(table)
        story.append(Spacer(1, 4))

    doc.build(story)
    return buffer.getvalue()


def _refs_names_only_for_game(g: dict) -> str:
    r1 = (g["slots"][1]["name"] or "").strip() or "—"
    r2 = (g["slots"][2]["name"] or "").strip() or "—"
    return f"{r1} / {r2}"


def build_referee_scorecards_pdf_bytes(selected_date: date) -> bytes:
    games = get_admin_print_rows_for_date(selected_date)

    buf = BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    page_w, page_h = A4

    outer_margin = 10 * mm
    gutter_x = 6 * mm
    gutter_y = 6 * mm

    card_w = (page_w - (2 * outer_margin) - gutter_x) / 2.0
    card_h = (page_h - (2 * outer_margin) - (2 * gutter_y)) / 3.0

    pad = 6 * mm

    TITLE_SIZE = 16
    TEAMS_MAX_SIZE = 13
    TEAMS_MIN_SIZE = 9
    REFS_SIZE = 11
    FIELD_TIME_SIZE = 12

    TEAM_ABOVE_NUM_MAX = 11
    TEAM_ABOVE_NUM_MIN = 8

    TRIES_NUM_SIZE = 18
    FOOT_LABEL_SIZE = 10

    WRITE_LINE_W = int(22 * 1.25)
    WRITE_LINE_THICK = 1.2

    def fit_bold_font_size(text: str, max_width: float, start_size: int, min_size: int) -> int:
        size = start_size
        while size > min_size:
            if c.stringWidth(text, "Helvetica-Bold", size) <= max_width:
                return size
            size -= 1
        return min_size

    def fit_left_text_size(text: str, max_width: float, start_size: int, min_size: int) -> int:
        size = start_size
        while size > min_size:
            if c.stringWidth(text, "Helvetica-Bold", size) <= max_width:
                return size
            size -= 1
        return min_size

    def draw_card(x0: float, y0: float, g: dict):
        c.setLineWidth(1)
        c.rect(x0, y0, card_w, card_h)

        left = x0 + pad
        right = x0 + card_w - pad
        y_top = y0 + card_h - pad
        max_text_w = right - left
        cx = x0 + card_w / 2.0

        c.setFont("Helvetica-Bold", TITLE_SIZE)
        c.drawCentredString(cx, y_top, "REFEREE SCORECARD")

        c.setLineWidth(0.8)
        c.line(left, y_top - 8, right, y_top - 8)

        teams_line = f"{g['home_team']} vs {g['away_team']}"
        teams_size = fit_bold_font_size(teams_line, max_text_w, TEAMS_MAX_SIZE, TEAMS_MIN_SIZE)
        c.setFont("Helvetica-Bold", teams_size)
        c.drawCentredString(cx, y_top - 26, teams_line)

        refs_line = _refs_names_only_for_game(g)
        c.setFont("Helvetica", REFS_SIZE)
        c.drawCentredString(cx, y_top - 44, refs_line)

        dt = dtparser.parse(g["start_dt"])
        field_time = f"{g['field_name']} @ {_time_12h(dt)}"
        c.setFont("Helvetica", FIELD_TIME_SIZE)
        c.drawCentredString(cx, y_top - 62, field_time)

        field_div_y = (y_top - 62) - 10
        c.setLineWidth(0.8)
        c.line(left, field_div_y, right, field_div_y)

        nums_left = left
        nums_right = right
        nums_span = nums_right - nums_left
        step = nums_span / 10.0

        wld_text = "W  /  L  /  D"
        wld_w = c.stringWidth(wld_text, "Helvetica-Bold", FOOT_LABEL_SIZE)

        team1_name_y = field_div_y - 16
        team1_nums_y = team1_name_y - 20

        INTER_TEAM_GAP = int(18 * 2.0)

        team2_name_y = team1_nums_y - INTER_TEAM_GAP
        team2_nums_y = team2_name_y - 20

        def draw_team_name_with_wld(team_name: str, y: float):
            nm = str(team_name)
            max_name_w = max_text_w - (wld_w + 6)
            size = fit_left_text_size(nm, max_name_w, TEAM_ABOVE_NUM_MAX, TEAM_ABOVE_NUM_MIN)

            c.setFont("Helvetica-Bold", size)
            c.drawString(left, y, nm)

            c.setFont("Helvetica-Bold", FOOT_LABEL_SIZE)
            c.drawRightString(right, y, wld_text)

        def draw_nums_row(y: float):
            c.setFont("Helvetica-Bold", TRIES_NUM_SIZE)
            for i in range(10):
                n = str(i + 1)
                x = nums_left + (step * (i + 0.5))
                c.drawCentredString(x, y, n)

        draw_team_name_with_wld(g["home_team"], team1_name_y)
        draw_nums_row(team1_nums_y)

        draw_team_name_with_wld(g["away_team"], team2_name_y)
        draw_nums_row(team2_nums_y)

        line_y = team2_nums_y - 14
        c.setLineWidth(0.8)
        c.line(left, line_y, right, line_y)

        line_x2 = right
        line_x1 = right - WRITE_LINE_W

        conduct_y = line_y - 18
        c.setFont("Helvetica-Bold", FOOT_LABEL_SIZE)
        c.drawString(left, conduct_y, "Conduct (/10)")
        c.setLineWidth(WRITE_LINE_THICK)
        c.line(line_x1, conduct_y - 3, line_x2, conduct_y - 3)

        unstrip_y = conduct_y - 22
        c.setFont("Helvetica-Bold", FOOT_LABEL_SIZE)
        c.drawString(left, unstrip_y, "Unstripped Players")
        c.setLineWidth(WRITE_LINE_THICK)
        c.line(line_x1, unstrip_y - 3, line_x2, unstrip_y - 3)

    for idx, g in enumerate(games):
        if idx > 0 and idx % 6 == 0:
            c.showPage()

        pos = idx % 6
        r = pos // 2
        col = pos % 2

        x0 = outer_margin + col * (card_w + gutter_x)
        y0 = page_h - outer_margin - (r + 1) * card_h - r * gutter_y

        draw_card(x0, y0, g)

    c.save()
    return buf.getvalue()


# ============================================================
# Printable XLSX helpers
# ============================================================
def build_admin_summary_xlsx_bytes(selected_date: date) -> bytes:
    games = get_admin_print_rows_for_date(selected_date)

    wb = Workbook()
    ws = wb.active
    ws.title = "Game Summary"

    # Styles
    title_font = Font(bold=True, size=16)
    header_font = Font(bold=True, size=11)
    bold_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    header_fill = PatternFill("solid", fgColor="D9D9D9")  # light grey
    group_fill = PatternFill("solid", fgColor="F2F2F2")

    # Title row
    ws["A1"] = "Game Summary"
    ws["A1"].font = title_font
    ws["A2"] = selected_date.isoformat()
    ws["A2"].font = bold_font

    row = 4

    if not games:
        ws[f"A{row}"] = "No games found for this date."
        out = BytesIO()
        wb.save(out)
        return out.getvalue()

    # Group by start time (same as PDF)
    grouped: dict[str, list[tuple[datetime, dict]]] = {}
    for g in games:
        dt = dtparser.parse(g["start_dt"])
        key = _time_12h(dt)
        grouped.setdefault(key, []).append((dt, g))

    group_keys = sorted(grouped.keys(), key=lambda k: min(dt for dt, _ in grouped[k]))

    # Column headings
    cols = ["Teams", "Field", "Start", "Referees"]

    for time_key in group_keys:
        # Group header
        ws[f"A{row}"] = f"Start time: {time_key}"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        ws[f"A{row}"].fill = group_fill
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(cols))
        row += 1

        # Table header
        for c, name in enumerate(cols, start=1):
            cell = ws.cell(row=row, column=c, value=name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
        row += 1

        # Rows
        for dt, g in grouped[time_key]:
            teams = f"{g['home_team']} vs {g['away_team']}"
            field = g["field_name"]
            start_str = _time_12h(dt)

            r1 = _format_ref_name(g["slots"][1]["name"], g["slots"][1]["status"])
            r2 = _format_ref_name(g["slots"][2]["name"], g["slots"][2]["status"])
            refs = f"{r1} / {r2}"

            ws.cell(row=row, column=1, value=teams).alignment = left
            ws.cell(row=row, column=2, value=field).alignment = left
            ws.cell(row=row, column=3, value=start_str).alignment = center
            ws.cell(row=row, column=4, value=refs).alignment = left
            row += 1

        row += 1  # spacer line

    # Column widths
    widths = [48, 16, 12, 34]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze panes at first header area (nice UX)
    ws.freeze_panes = "A4"

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


# ============================================================
# Offers
# ============================================================
# ============================================================
# Offers (bulk helpers)
# ============================================================

def _offers_for_date_rows(selected_date: date) -> list[sqlite3.Row]:
    """
    Returns all assignment slots for games on selected_date, including referee + game details.
    """
    start_min = datetime.combine(selected_date, datetime.min.time()).isoformat(timespec="seconds")
    start_max = datetime.combine(selected_date + timedelta(days=1), datetime.min.time()).isoformat(timespec="seconds")

    conn = db()
    rows = conn.execute(
        """
        SELECT
            a.id AS assignment_id,
            a.status AS status,
            a.referee_id AS referee_id,

            r.name  AS ref_name,
            r.email AS ref_email,

            g.id AS game_id,
            g.home_team,
            g.away_team,
            g.field_name,
            g.start_dt
        FROM games g
        JOIN assignments a ON a.game_id = g.id
        LEFT JOIN referees r ON r.id = a.referee_id
        WHERE g.start_dt >= ? AND g.start_dt < ?
        ORDER BY g.start_dt ASC, g.field_name ASC, g.home_team ASC, a.slot_no ASC
        """,
        (start_min, start_max),
    ).fetchall()
    conn.close()
    return rows


def count_bulk_offer_candidates(selected_date: date) -> dict:
    """
    Counts what would happen if we ran bulk-offer for the selected date.
    """
    rows = _offers_for_date_rows(selected_date)

    counts = {
        "candidates": 0,
        "skipped_no_ref": 0,
        "skipped_blackout": 0,
        "skipped_already_offered": 0,
        "skipped_confirmed": 0,
        "skipped_other": 0,
    }

    for r in rows:
        status = (r["status"] or "").strip().upper()
        ref_id = r["referee_id"]

        if ref_id is None:
            counts["skipped_no_ref"] += 1
            continue

        if status in ("ACCEPTED", "ASSIGNED"):
            counts["skipped_confirmed"] += 1
            continue

        if status == "OFFERED":
            counts["skipped_already_offered"] += 1
            continue

        # We only bulk-send for NOT_OFFERED (your “pre-allocated but not sent yet” state)
        if status != "NOT_OFFERED":
            counts["skipped_other"] += 1
            continue

        gdate = dtparser.parse(r["start_dt"]).date()
        if referee_has_blackout(int(ref_id), gdate):
            counts["skipped_blackout"] += 1
            continue

        counts["candidates"] += 1

    return counts


def send_bulk_offers_for_date(selected_date: date) -> dict:
    """
    Sends offers for ALL NOT_OFFERED assignments on selected_date (skipping blackouts/confirmed/already offered).
    Returns a summary dict (sent / skipped / failed).
    """
    rows = _offers_for_date_rows(selected_date)

    summary = {
        "sent": 0,
        "failed": 0,
        "skipped_no_ref": 0,
        "skipped_blackout": 0,
        "skipped_already_offered": 0,
        "skipped_confirmed": 0,
        "skipped_other": 0,
        "failures": [],  # list of strings
    }

    for r in rows:
        assignment_id = int(r["assignment_id"])
        status = (r["status"] or "").strip().upper()
        ref_id = r["referee_id"]

        if ref_id is None:
            summary["skipped_no_ref"] += 1
            continue

        if status in ("ACCEPTED", "ASSIGNED"):
            summary["skipped_confirmed"] += 1
            continue

        if status == "OFFERED":
            summary["skipped_already_offered"] += 1
            continue

        if status != "NOT_OFFERED":
            summary["skipped_other"] += 1
            continue

        g_start_dt = dtparser.parse(r["start_dt"])
        gdate = g_start_dt.date()

        if referee_has_blackout(int(ref_id), gdate):
            summary["skipped_blackout"] += 1
            continue

        # Build a game dict compatible with send_offer_email_and_mark_offered()
        game_row = {
            "home_team": r["home_team"],
            "away_team": r["away_team"],
            "field_name": r["field_name"],
        }

        referee_name = (r["ref_name"] or "").strip() or "Referee"
        referee_email = (r["ref_email"] or "").strip()

        if not referee_email:
            summary["failed"] += 1
            summary["failures"].append(
                f"{referee_name} — missing email (assignment_id={assignment_id})"
            )
            continue

        # IMPORTANT: avoid multiple live tokens for the same assignment
        # If an offer already exists (e.g. admin clicked earlier then changed mind),
        # clear any old ones before creating a new offer token.
        conn = db()
        try:
            _delete_offers_for_assignment(conn, assignment_id)
            conn.commit()
        finally:
            conn.close()

        # Use your existing sending logic (creates offer token, emails, marks OFFERED, rolls back token on failure)
        msg_key = f"bulk_offer_msg_{assignment_id}"
        send_offer_email_and_mark_offered(
            assignment_id=assignment_id,
            referee_name=referee_name,
            referee_email=referee_email,
            game=game_row,
            start_dt=g_start_dt,
            msg_key=msg_key,
        )

        # Determine success/fail from the msg_key that function sets
        msg = st.session_state.get(msg_key, "")
        if msg.startswith("Offer emailed successfully"):
            summary["sent"] += 1
        else:
            summary["failed"] += 1
            summary["failures"].append(
                f"{referee_name} — {r['home_team']} vs {r['away_team']} @ {r['field_name']} — {msg}"
            )

    return summary


def create_offer(assignment_id: int) -> str:
    token = secrets.token_urlsafe(24)
    conn = db()
    conn.execute(
        """
        INSERT INTO offers(assignment_id, token, created_at)
        VALUES (?, ?, ?)
        """,
        (assignment_id, token, now_iso()),
    )
    conn.commit()
    conn.close()
    return token


def delete_offer_by_token(token: str):
    conn = db()
    conn.execute("DELETE FROM offers WHERE token=?", (token,))
    conn.commit()
    conn.close()


def resolve_offer(token: str, response: str) -> tuple[bool, str]:
    response = (response or "").strip().upper()
    if response not in ("ACCEPTED", "DECLINED"):
        return False, "Invalid response."

    conn = db()
    offer = conn.execute(
        """
        SELECT id, assignment_id
        FROM offers
        WHERE token=?
        """,
        (token,),
    ).fetchone()

    if not offer:
        conn.close()
        return False, "Invalid or unknown offer link."

    conn.execute(
        """
        UPDATE offers
        SET responded_at=?, response=?
        WHERE id=?
        """,
        (now_iso(), response, offer["id"]),
    )

    new_status = "ACCEPTED" if response == "ACCEPTED" else "DECLINED"
    conn.execute(
        """
        UPDATE assignments
        SET status=?, updated_at=?
        WHERE id=?
        """,
        (new_status, now_iso(), offer["assignment_id"]),
    )

    conn.commit()
    conn.close()
    return True, f"Thanks — you have {response.lower()} the offer."


def send_offer_email_and_mark_offered(
    *,
    assignment_id: int,
    referee_name: str,
    referee_email: str,
    game,
    start_dt,
    msg_key: str,
):
    token = create_offer(assignment_id)

    try:
        cfg = smtp_settings()
        base = cfg.get("app_base_url", "").rstrip("/")
        if not base:
            raise RuntimeError("APP_BASE_URL is missing. Add it in Render environment variables.")

        game_line = f"{game['home_team']} vs {game['away_team']}"
        when_line = start_dt.strftime("%Y-%m-%d %I:%M %p").lstrip("0")
        subject = f"{referee_name} — Match assignment: {game_line}"

        portal_url = f"{base}/?offer_token={token}"

        text = (
            f"Hi {referee_name},\n\n"
            f"You have a match assignment offer:\n"
            f"- Game: {game_line}\n"
            f"- Field: {game['field_name']}\n"
            f"- Start: {when_line}\n\n"
            f"View and respond here:\n{portal_url}\n"
        )

        html = f"""
        <div style="font-family: Arial, sans-serif; line-height:1.4;">
          <p>Hi {referee_name},</p>
          <p>You have a match assignment offer:</p>
          <ul>
            <li><b>Game:</b> {game_line}</li>
            <li><b>Field:</b> {game['field_name']}</li>
            <li><b>Start:</b> {when_line}</li>
          </ul>
          <p>
            <a href="{portal_url}" style="display:inline-block;padding:10px 14px;background:#1565c0;color:#fff;text-decoration:none;border-radius:6px;">
              View offer
            </a>
          </p>
          <p style="color:#666;font-size:12px;">
            If the button doesn’t work, copy and paste this link:<br>{portal_url}
          </p>
        </div>
        """

        send_html_email(referee_email, referee_name, subject, html, text_body=text)

        set_assignment_status(assignment_id, "OFFERED")
        st.session_state[msg_key] = "Offer emailed successfully and marked as OFFERED."
    except Exception as e:
        delete_offer_by_token(token)
        st.session_state[msg_key] = f"Email failed — offer not created: {e}"


# ============================================================
# Referee portal (My Offers)
# ============================================================
def get_offer_details_by_token(token: str):
    conn = db()
    row = conn.execute(
        """
        SELECT
            o.id AS offer_id,
            o.token,
            o.created_at,
            o.responded_at,
            o.response,
            a.id AS assignment_id,
            a.slot_no,
            a.referee_id,
            a.status,
            r.name AS ref_name,
            r.email AS ref_email,
            g.home_team,
            g.away_team,
            g.field_name,
            g.start_dt
        FROM offers o
        JOIN assignments a ON a.id = o.assignment_id
        JOIN games g ON g.id = a.game_id
        LEFT JOIN referees r ON r.id = a.referee_id
        WHERE o.token=?
        LIMIT 1
        """,
        (token,),
    ).fetchone()
    conn.close()
    return row


def list_offers_for_referee(referee_id: int):
    conn = db()
    rows = conn.execute(
        """
        SELECT
            o.id AS offer_id,
            o.token,
            o.created_at,
            o.responded_at,
            o.response,
            a.id AS assignment_id,
            a.slot_no,
            a.status,
            g.home_team,
            g.away_team,
            g.field_name,
            g.start_dt
        FROM offers o
        JOIN assignments a ON a.id = o.assignment_id
        JOIN games g ON g.id = a.game_id
        WHERE a.referee_id=?
        ORDER BY o.created_at DESC
        """,
        (referee_id,),
    ).fetchall()
    conn.close()
    return rows


def maybe_handle_referee_portal_login():
    if not REF_PORTAL_ENABLED:
        return

    qp = st.query_params
    offer_token = qp.get("offer_token")
    if not offer_token:
        return

    offer = get_offer_details_by_token(offer_token)
    if not offer:
        st.title("My Offers")
        st.error("That offer link is invalid.")
        st.stop()

    if offer["referee_id"] is None:
        st.title("My Offers")
        st.error("This offer is not linked to a referee. Please contact the administrator.")
        st.stop()

    st.session_state["referee_id"] = int(offer["referee_id"])
    st.session_state["referee_name"] = offer["ref_name"] or "Referee"
    st.session_state["referee_email"] = offer["ref_email"] or ""

    st.query_params.pop("offer_token", None)
    st.rerun()


def referee_logout_button():
    if st.session_state.get("referee_id"):
        c1, c2 = st.columns([3, 1])
        with c1:
            st.caption(
                f"Logged in as: {st.session_state.get('referee_name')} "
                f"({st.session_state.get('referee_email')})"
            )
        with c2:
            if st.button("Log out", key="ref_logout_btn"):
                st.session_state.pop("referee_id", None)
                st.session_state.pop("referee_name", None)
                st.session_state.pop("referee_email", None)
                st.rerun()


def render_my_offers_page() -> bool:
    if not REF_PORTAL_ENABLED:
        return False

    ref_id = st.session_state.get("referee_id")
    if not ref_id:
        return False

    st.title("My Offers")
    referee_logout_button()
    st.markdown("---")

    offers = list_offers_for_referee(int(ref_id))
    if not offers:
        st.info("You have no offers at the moment.")
        return True

    for o in offers:
        start_dt = dtparser.parse(o["start_dt"])
        title = f"{o['home_team']} vs {o['away_team']}"
        subtitle = f"Field: {o['field_name']} • Start: {_time_12h(start_dt)} • Slot {o['slot_no']}"

        with st.container(border=True):
            st.subheader(title)
            st.caption(subtitle)

            current_resp = (o["response"] or "").strip().upper()
            if o["responded_at"] and current_resp:
                if current_resp == "DECLINED":
                    st.markdown(
                        "<div style='font-weight:700;color:#c62828;'>Response recorded: DECLINED</div>",
                        unsafe_allow_html=True,
                    )
                elif current_resp == "ACCEPTED":
                    st.markdown(
                        "<div style='font-weight:700;color:#2e7d32;'>Response recorded: ACCEPTED</div>",
                        unsafe_allow_html=True,
                    )
                else:
                    st.info(f"Response recorded: {current_resp}")

                st.caption("You can change your response below if needed.")

            c1, c2 = st.columns(2)

            if c1.button("Accept", key=f"portal_acc_{o['token']}"):
                ok, msg = resolve_offer(o["token"], "ACCEPTED")
                if ok:
                    st.success("Accepted. Thank you.")
                else:
                    st.error(msg)
                st.rerun()

            if c2.button("Decline", key=f"portal_dec_{o['token']}"):
                ok, msg = resolve_offer(o["token"], "DECLINED")
                if ok:
                    st.success("Declined. Thank you.")
                else:
                    st.error(msg)
                st.rerun()

    return True


def maybe_handle_offer_response():
    qp = st.query_params
    token = qp.get("token")
    action = qp.get("action")
    if token and action in ("accept", "decline"):
        response = "ACCEPTED" if action == "accept" else "DECLINED"
        ok, msg = resolve_offer(token, response)
        st.title("Referee Response")
        if ok:
            st.success(msg)
        else:
            st.error(msg)
        st.info("You can close this page now.")
        st.stop()

# ============================================================
# APP START
# ============================================================
init_db()
ensure_meta_table()
maybe_auto_backup()

maybe_handle_referee_portal_login()
maybe_handle_offer_response()

if render_my_offers_page():
    st.stop()

handle_admin_login_via_query_params()
maybe_restore_admin_from_session_param()

st.title("Referee Allocator — MVP")
st.caption(f"Database file: {DB_PATH}")

# Bootstrap: create first admin if none exist
if admin_count() == 0:
    st.warning("Initial setup: No administrators exist yet.")
    st.write("Enter your email to create the first admin account (one-time setup).")
    first_email = st.text_input("Your admin email", key="first_admin_email")
    if st.button("Create first admin", key="create_first_admin_btn"):
        if not first_email.strip():
            st.error("Please enter an email.")
        else:
            add_admin(first_email)
            st.success("First admin created. Now request a login link below.")
    st.stop()

# Login screen
if not st.session_state.get("admin_email"):
    st.subheader("Admin Login")
    st.write("Enter your email to log in.")

    email = st.text_input("Admin email", key="login_email")
    typed_email = (email or "").strip().lower()

    # ------------------------------------------------------------
    # SIMPLE BYPASS LOGIN
    # Normal admins: instant login if in admin allowlist
    # Super admin: must use email/token
    # ------------------------------------------------------------
    if typed_email and (not is_super_admin_email(typed_email)) and is_admin_email_allowed(typed_email):
        st.success("Recognised administrator — logging you in...")
        st.session_state["admin_email"] = typed_email
        st.rerun()

    st.markdown("---")

    # ------------------------------------------------------------
    # Super Admin login (email token)
    # ------------------------------------------------------------
    if is_super_admin_email(typed_email):
        st.info("Super admin — use login link.")
        if st.button("Send login link", key="send_login_link_btn"):
            try:
                send_admin_login_email(typed_email)
                st.success("Login link sent. Check your email.")
            except Exception as e:
                st.error(str(e))
    else:
        st.caption("If you can’t log in, ask Landon to add you as admin.")

    # ------------------------------------------------------------
    # Emergency Admin Link (optional)
    # ------------------------------------------------------------
    if os.getenv("SHOW_ADMIN_LINK", "false").lower() == "true":
        st.markdown("---")
        st.subheader("Emergency Admin Link")

        if st.button("Generate admin login link", key="show_admin_link_btn"):
            if not typed_email:
                st.error("Enter email first.")
            elif not is_admin_email_allowed(typed_email):
                st.error("Not an authorised admin.")
            else:
                cfg = smtp_settings()
                base = (cfg.get("app_base_url") or "").rstrip("/")
                token = create_admin_login_token(typed_email, minutes_valid=15)
                login_url = f"{base}/?admin_login=1&token={token}"
                st.code(login_url)

    st.stop()


    # ------------------------------------------------------------
    # DEV Admin URL (no email) — enabled via environment variable
    # ------------------------------------------------------------
    if os.getenv("DEV_ADMIN_URL_ENABLED", "false").lower() == "true":
        st.markdown("---")
        st.subheader("DEV Admin Login (no email)")

        if st.button("Generate DEV admin login URL", key="dev_admin_url_btn"):
            dev_email = os.getenv("DEV_ADMIN_EMAIL", "").strip().lower()

            if not dev_email:
                st.error("DEV_ADMIN_EMAIL is not set in environment variables.")
            elif not is_admin_email_allowed(dev_email):
                st.error(f"{dev_email} is not an active admin.")
            else:
                cfg = smtp_settings()
                base = (cfg.get("app_base_url") or "").rstrip("/")

                if not base:
                    st.error("APP_BASE_URL is missing.")
                else:
                    # 90-day expiry (adjust if needed)
                    expires_at = (
                        datetime.now(timezone.utc) + timedelta(days=90)
                    ).isoformat(timespec="seconds")

                    token = create_admin_session_with_expires_at(dev_email, expires_at)
                    url = f"{base}/?session={token}"

                    st.success("DEV admin login URL created:")
                    st.code(url)
                    st.caption(
                        "Bookmark this URL. Disable DEV_ADMIN_URL_ENABLED when finished."
                    )


    st.stop()

# Logged in view
admin_logout_button()

tabs = st.tabs(["Admin", "Ladder", "Import", "Blackouts", "Administrators"])

# ============================================================
# Admin tab
# ============================================================
with tabs[0]:
    st.subheader("Games & Assignments")

    auto = st.toggle("Auto-refresh every 5 seconds", value=True, key="auto_refresh_toggle")
    if auto:
        st_autorefresh(interval=5000, key="auto_refresh_tick")

    if st.button("Refresh status", key="refresh_status_btn"):
        st.rerun()

    games = get_games()
    refs = get_referees()

    if not games:
        st.info("Import a Games CSV first (Import tab).")
        st.stop()

    all_dates = sorted({game_local_date(g) for g in games})
    today = date.today()
    default_idx = 0
    for i, d in enumerate(all_dates):
        if d >= today:
            default_idx = i
            break

    # ------------------------------------------------------------
    # Show games for date (display as dd-MMM-yy)
    # Use string labels and map back to date objects.
    # Also: clear stale session_state values from older versions.
    # ------------------------------------------------------------
    date_label_to_date = {d.strftime("%d-%b-%y"): d for d in all_dates}
    date_labels = list(date_label_to_date.keys())

    # ✅ clear stale old value (e.g. "2026-02-04") that is no longer a valid option
    prev = st.session_state.get("admin_show_games_for_date")
    if prev and prev not in date_labels:
        st.session_state.pop("admin_show_games_for_date", None)

    default_label = all_dates[default_idx].strftime("%d-%b-%y")

    selected_label = st.selectbox(
        "Show games for date",
        date_labels,
        index=date_labels.index(default_label),
        key="admin_show_games_for_date",
    )

    selected_date = date_label_to_date[selected_label]
    date_key = selected_date.isoformat()

    # Keep scroll position stable across auto-refresh/reruns (per selected date)
    preserve_scroll(scroll_key=f"refalloc_admin_scroll_{selected_date.isoformat()}")

    # ------------------------------------------------------------
    # BULK OFFERS — push all NOT_OFFERED slots for this date
    # ------------------------------------------------------------
    st.markdown("---")
    st.subheader("Bulk offers")

    bulk_counts = count_bulk_offer_candidates(selected_date)

    c_bulk1, c_bulk2 = st.columns([2, 1], gap="large")

    with c_bulk1:
        st.caption(
            f"Candidates to send now: {bulk_counts['candidates']}  •  "
            f"Skipped: blackout={bulk_counts['skipped_blackout']}, "
            f"already offered={bulk_counts['skipped_already_offered']}, "
            f"confirmed={bulk_counts['skipped_confirmed']}"
        )

        confirm_bulk = st.checkbox(
            "Yes — send OFFERS for ALL candidates on this date",
            value=False,
            key=f"confirm_bulk_offers_{selected_date.isoformat()}",
        )

        if st.button(
            "Send OFFERS for all games on this date",
            key=f"send_bulk_offers_btn_{selected_date.isoformat()}",
            disabled=(not confirm_bulk or bulk_counts["candidates"] == 0),
        ):
            result = send_bulk_offers_for_date(selected_date)
            st.session_state[f"bulk_offer_result_{selected_date.isoformat()}"] = result
            st.rerun()

    with c_bulk2:
        res = st.session_state.get(f"bulk_offer_result_{selected_date.isoformat()}")
        if res:
            status_badge(f"Sent: {res['sent']}", bg="#2e7d32")
            st.caption(
                f"Failed: {res['failed']} • "
                f"Skipped: blackout={res['skipped_blackout']}, "
                f"already offered={res['skipped_already_offered']}, "
                f"confirmed={res['skipped_confirmed']}"
            )
            if res.get("failures"):
                with st.expander("Show failures"):
                    for f in res["failures"][:50]:
                        st.write(f"- {f}")
                    if len(res["failures"]) > 50:
                        st.caption(f"(Showing first 50 of {len(res['failures'])})")

    count_games = sum(1 for g in games if game_local_date(g) == selected_date)
    st.caption(f"{count_games} game(s) on {selected_date.isoformat()}")

    week_start, week_end_excl = iso_week_window(selected_date)

    main_col, side_col = st.columns([3, 1], gap="large")

    with side_col:
        st.markdown("### Referee workload")
        st.caption("All-time accepted/assigned (all games)")
        df_work = get_referee_workload_all_time()

        if df_work.empty:
            st.info("No referees found.")
        else:
            st.dataframe(df_work, use_container_width=True, hide_index=True)
            total_acc = int(df_work["Accepted"].sum()) if "Accepted" in df_work.columns else 0
            st.caption(f"Total accepted/assigned slots (all-time): {total_acc}")

    with main_col:
        accepted_slots, total_slots = get_acceptance_progress_for_window(week_start, week_end_excl)

        pct = (accepted_slots / total_slots) if total_slots else 0.0
        pct_clamped = max(0.0, min(1.0, pct))

        if total_slots == 0:
            bar_color = "#9e9e9e"
        elif pct_clamped < 0.50:
            bar_color = "#c62828"
        elif pct_clamped < 0.90:
            bar_color = "#ffb300"
        else:
            bar_color = "#2e7d32"

        not_accepted_names = list_referees_not_accepted_for_window(week_start, week_end_excl)

        c_bar, c_list = st.columns([1, 2])


        with c_bar:
            height_px = 24
            st.markdown(
                f"""
                <div style="font-size:12px; color:#666; margin-bottom:6px;">
                  <b>Week acceptance (ISO)</b> — {accepted_slots}/{total_slots}
                </div>

                <div style="width:100%; background:#e0e0e0; border-radius:{height_px}px; height:{height_px}px; overflow:hidden;">
                  <div style="width:{pct_clamped*100:.1f}%; background:{bar_color}; height:{height_px}px;"></div>
                </div>
                """,
                unsafe_allow_html=True,
            )

        with c_list:
            has_offers = has_any_offers_for_window(week_start, week_end_excl)

            if not has_offers:
                st.caption("")
            else:
                st.markdown(
                    "<div style='font-size:12px; color:#666; margin-bottom:6px;'><b>Yet to ACCEPT (unique)</b></div>",
                    unsafe_allow_html=True,
                )

                if not not_accepted_names:
                    st.markdown(
                        "<div style='font-size:12px; color:#2e7d32;'>All accepted ✅</div>",
                        unsafe_allow_html=True,
                    )
                else:
                    items_html = ", ".join([f"<span>{n}</span>" for n in not_accepted_names])
                    st.markdown(
                        f"""
                        <div style="
                            font-size:12px;
                            color:#ffb300;
                            line-height:1.6;
                            display:flex;
                            flex-wrap:wrap;
                            gap:6px;
                            align-items:center;
                        ">
                          {items_html}
                        </div>
                        """,
                        unsafe_allow_html=True,
                    )

        st.markdown("---")
        st.subheader("Printable Summary")

        c_pdf1, c_pdf2, c_x1, c_x2, c_pdf3, c_pdf4 = st.columns([1, 2, 1, 2, 1, 2])

        date_key = selected_date.isoformat()

        # =========================
        # SUMMARY PDF
        # =========================
        with c_pdf1:
            if st.button("Build Summary PDF", key=f"build_pdf_btn_{date_key}"):
                try:
                    pdf_bytes = build_admin_summary_pdf_bytes(selected_date)
                    st.session_state[f"admin_summary_pdf_bytes_{date_key}"] = pdf_bytes
                    st.success("Summary PDF built.")
                except Exception as e:
                    st.error(f"Failed to build Summary PDF: {e}")

        with c_pdf2:
            pdf_bytes = st.session_state.get(f"admin_summary_pdf_bytes_{date_key}")
            if pdf_bytes:
                st.download_button(
                    "Download Summary PDF",
                    data=pdf_bytes,
                    file_name=f"game_summary_{date_key}.pdf",
                    mime="application/pdf",
                    key=f"download_pdf_btn_{date_key}",
                )
            else:
                st.caption("Click **Build Summary PDF** first.")

        # =========================
        # SUMMARY XLSX
        # =========================
        with c_x1:
            if st.button("Build Summary XLSX", key=f"build_xlsx_btn_{date_key}"):
                try:
                    xlsx_bytes = build_admin_summary_xlsx_bytes(selected_date)
                    st.session_state[f"admin_summary_xlsx_bytes_{date_key}"] = xlsx_bytes
                    st.success("Summary XLSX built.")
                except Exception as e:
                    st.error(f"Failed to build Summary XLSX: {e}")

        with c_x2:
            xlsx_bytes = st.session_state.get(f"admin_summary_xlsx_bytes_{date_key}")
            if xlsx_bytes:
                st.download_button(
                    "Download for Excel / Google Sheets",
                    data=xlsx_bytes,
                    file_name=f"game_summary_{date_key}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"download_xlsx_btn_{date_key}",
                )
                st.caption("Tip: Upload to Google Sheets — opens natively.")
            else:
                st.caption("Click **Build Summary XLSX** first.")

        # =========================
        # SCORECARDS PDF
        # =========================
        with c_pdf3:
            if st.button("Build Referee Scorecards", key=f"build_scorecards_btn_{date_key}"):
                try:
                    sc_bytes = build_referee_scorecards_pdf_bytes(selected_date)
                    st.session_state[f"ref_scorecards_pdf_bytes_{date_key}"] = sc_bytes
                    st.success("Scorecards PDF built.")
                except Exception as e:
                    st.error(f"Failed to build Scorecards PDF: {e}")

        with c_pdf4:
            sc_bytes = st.session_state.get(f"ref_scorecards_pdf_bytes_{date_key}")
            if sc_bytes:
                st.download_button(
                    "Download Scorecards PDF",
                    data=sc_bytes,
                    file_name=f"referee_scorecards_{date_key}.pdf",
                    mime="application/pdf",
                    key=f"download_scorecards_btn_{date_key}",
                )
            else:
                st.caption("Click **Build Referee Scorecards** first.")


        for g in games:
            if game_local_date(g) != selected_date:
                continue

            start_dt = dtparser.parse(g["start_dt"])
            gdate = game_local_date(g)

            ref_options = ["— Select referee —"]
            ref_lookup = {}
            for r in refs:
                label = f"{r['name']} ({r['email']})"
                if referee_has_blackout(r["id"], gdate):
                    label = f"🚫 {label} — blackout"
                ref_options.append(label)
                ref_lookup[label] = r["id"]

            with st.container(border=True):
                st.markdown(
                    f"**{g['home_team']} vs {g['away_team']}**  \n"
                    f"Field: **{g['field_name']}**  \n"
                    f"Start: **{start_dt.strftime('%Y-%m-%d %I:%M %p').lstrip('0')}**"
                )

                assigns = get_assignments_for_game(g["id"])
                cols = st.columns(2)

                for col_idx, a in enumerate(assigns):
                    with cols[col_idx]:
                        st.markdown(f"#### Slot {a['slot_no']}")

                        status = (a["status"] or "").strip().upper()
                        st.caption(f"assignment_id={a['id']} | status={status} | updated_at={a['updated_at']}")

                        current_ref_label = None
                        if a["referee_id"] is not None and a["ref_name"] and a["ref_email"]:
                            current_ref_label = f"{a['ref_name']} ({a['ref_email']})"
                            if referee_has_blackout(a["referee_id"], gdate):
                                current_ref_label = f"🚫 {current_ref_label} — blackout"

                        default_index = 0
                        if current_ref_label and current_ref_label in ref_options:
                            default_index = ref_options.index(current_ref_label)

                        refpick_key = f"refpick_{g['id']}_{a['slot_no']}"
                        pick = st.selectbox(
                            "Referee",
                            ref_options,
                            index=default_index,
                            key=refpick_key,
                            disabled=(status in ("ACCEPTED", "ASSIGNED")),
                        )

                        if pick != "— Select referee —":
                            chosen_ref_id = ref_lookup[pick]
                            if status in ("ACCEPTED", "ASSIGNED"):
                                st.info("This slot is locked (ACCEPTED/ASSIGNED). Use Action → RESET to change it.")
                            else:
                                if a["referee_id"] != chosen_ref_id:
                                    set_assignment_ref(a["id"], chosen_ref_id)
                                    st.rerun()
                        else:
                            if a["referee_id"] is not None:
                                clear_assignment(a["id"])
                                st.session_state[refpick_key] = "— Select referee —"
                                st.rerun()

                        blackout = False
                        if a["referee_id"] is not None:
                            blackout = referee_has_blackout(a["referee_id"], gdate)

                        if status == "ACCEPTED":
                            status_badge(f"✅ {a['ref_name']} — ACCEPTED", bg="#2e7d32")
                        elif status == "ASSIGNED":
                            status_badge(f"✅ {a['ref_name']} — ASSIGNED", bg="#2e7d32")
                        elif status == "DECLINED":
                            status_badge(f"❌ {a['ref_name']} — DECLINED", bg="#c62828")
                        elif status == "OFFERED":
                            status_badge(f"⬜ {a['ref_name']} — OFFERED", bg="#546e7a")
                        elif a["referee_id"] is not None:
                            status_badge(f"⬛ {a['ref_name']} — NOT OFFERED YET", bg="#424242")
                        else:
                            st.caption("EMPTY")

                        if blackout:
                            st.warning(f"Blackout date conflict: {gdate.isoformat()}")

                        action_key = f"action_{a['id']}"
                        msg_key = f"msg_{a['id']}"
                        st.session_state.setdefault(action_key, "—")

                        action_options = ["—", "OFFER", "ASSIGN", "DELETE", "RESET"]
                        if status in ("ACCEPTED", "ASSIGNED"):
                            action_options = ["—", "RESET", "DELETE"]

                        def on_action_change(
                            assignment_id=a["id"],
                            game_row=g,
                            start_dt=start_dt,
                            gdate=gdate,
                            action_key=action_key,
                            msg_key=msg_key,
                            refpick_key=refpick_key,
                        ):
                            choice = st.session_state.get(action_key, "—")
                            st.session_state.pop(msg_key, None)

                            if choice == "—":
                                return

                            live_a = get_assignment_live(assignment_id)
                            if not live_a:
                                st.session_state[msg_key] = "Could not load assignment."
                                st.session_state[action_key] = "—"
                                st.rerun()
                                return

                            live_ref_id = live_a["referee_id"]
                            live_ref_name = live_a["ref_name"]
                            live_ref_email = live_a["ref_email"]
                            live_status = (live_a["status"] or "").strip().upper()

                            live_blackout = False
                            if live_ref_id is not None:
                                live_blackout = referee_has_blackout(int(live_ref_id), gdate)

                            if live_ref_id is None and choice in ("OFFER", "ASSIGN"):
                                st.session_state[msg_key] = "Select a referee first."
                                st.session_state[action_key] = "—"
                                return

                            if choice == "OFFER" and live_status in ("ACCEPTED", "ASSIGNED"):
                                st.session_state[msg_key] = "This slot is already confirmed (ACCEPTED/ASSIGNED)."
                                st.session_state[action_key] = "—"
                                return

                            if choice == "OFFER":
                                if live_blackout:
                                    st.session_state[msg_key] = (
                                        "Offer blocked: referee is unavailable on this date (blackout). "
                                        "You can still ASSIGN manually if needed."
                                    )
                                    st.session_state[action_key] = "—"
                                    return

                                send_offer_email_and_mark_offered(
                                    assignment_id=assignment_id,
                                    referee_name=live_ref_name,
                                    referee_email=live_ref_email,
                                    game=game_row,
                                    start_dt=start_dt,
                                    msg_key=msg_key,
                                )

                            elif choice == "ASSIGN":
                                set_assignment_status(assignment_id, "ASSIGNED")
                                st.session_state[msg_key] = "Assigned."

                            elif choice in ("DELETE", "RESET"):
                                clear_assignment(assignment_id)
                                st.session_state[refpick_key] = "— Select referee —"
                                st.session_state[msg_key] = "Slot cleared (EMPTY)."

                            st.session_state[action_key] = "—"
                            st.rerun()

                        st.selectbox(
                            "Action",
                            action_options,
                            key=action_key,
                            on_change=on_action_change,
                        )

                        if st.session_state.get(msg_key):
                            st.info(st.session_state[msg_key])

# ============================================================
# LADDER HELPERS — Season ladder (Option A: AS AT DATE)
# ============================================================

def get_season_start_date() -> date | None:
    """
    Season start = earliest game date in the database.
    """
    conn = db()
    row = conn.execute(
        "SELECT MIN(start_dt) AS min_dt FROM games"
    ).fetchone()
    conn.close()

    if not row or not row["min_dt"]:
        return None

    try:
        return dtparser.parse(row["min_dt"]).date()
    except Exception:
        return None


def ladder_validation_warnings_for_date(d: date) -> list[str]:
    """
    Validates only games ON the selected date.
    """
    warnings: list[str] = []

    games = [g for g in get_games() if game_local_date(g) == d]

    for g in games:
        for t in (g["home_team"], g["away_team"]):
            if t and not get_team_division(t):
                warnings.append(f"Missing division for team: {t}")

        gr = get_game_result(int(g["id"]))
        if not gr:
            warnings.append(f"Missing result: {g['home_team']} vs {g['away_team']}")
            continue

        if gr["home_defaulted"] and gr["away_defaulted"]:
            warnings.append(
                f"Invalid DEFAULTED flags (both marked): "
                f"{g['home_team']} vs {g['away_team']}"
            )

    return warnings


def ladder_validation_warnings_as_at(as_at_date: date) -> list[str]:
    """
    Validates ALL games from season start → as_at_date.
    """
    warnings: list[str] = []

    season_start = get_season_start_date()
    if not season_start:
        return ["No games found (cannot determine season start)."]

    games = [
        g for g in get_games()
        if season_start <= game_local_date(g) <= as_at_date
    ]

    for g in games:
        for t in (g["home_team"], g["away_team"]):
            if t and not get_team_division(t):
                warnings.append(f"Missing division for team: {t}")

        gr = get_game_result(int(g["id"]))
        if not gr:
            warnings.append(
                f"Missing result: {g['home_team']} vs {g['away_team']} "
                f"({game_local_date(g)})"
            )
            continue

        if gr["home_defaulted"] and gr["away_defaulted"]:
            warnings.append(
                f"Invalid DEFAULTED flags (both marked): "
                f"{g['home_team']} vs {g['away_team']} "
                f"({game_local_date(g)})"
            )

    return warnings
# ============================================================
# Ladder tab
# ============================================================
with tabs[1]:
    st.subheader("Competition Ladder (Admin)")
    st.caption("Enter team divisions + opening balance + game results, then view ladder + audit breakdown for fault finding.")

    games = get_games()
    if not games:
        st.info("Import a Games CSV first.")
        st.stop()

    all_dates = sorted({game_local_date(g) for g in games})
    if not all_dates:
        st.info("No game dates found.")
        st.stop()

    today = date.today()
    default_idx = 0
    for i, d in enumerate(all_dates):
        if d >= today:
            default_idx = i
            break

    selected_date = st.selectbox(
        "Ladder date",
        all_dates,
        index=default_idx,
        key="ladder_date_select",
    )

    todays_games = [g for g in games if game_local_date(g) == selected_date]
    st.caption(f"{len(todays_games)} game(s) on {selected_date.isoformat()}")

    st.markdown("---")

    # ------------------------------------------------------------
    # Team divisions + opening balance (HIDDEN by default)
    # Auto-expands only when needed (missing divisions for today's teams)
    # ------------------------------------------------------------
    teams_today = sorted(
        {(g["home_team"] or "").strip() for g in todays_games}
        | {(g["away_team"] or "").strip() for g in todays_games}
    )
    teams_today = [t for t in teams_today if t]

    teams_rows = list_teams()
    existing = {
        r["name"]: {
            "division": (r["division"] or "").strip(),
            "opening": int(r["opening_balance"] or 0),
        }
        for r in teams_rows
    }

    if not teams_today:
        st.info("No teams found for this date.")
        st.stop()

    missing_div_teams = [t for t in teams_today if not (existing.get(t, {}).get("division") or "").strip()]

    # Small status line so the user knows where to go only if required
    if missing_div_teams:
        st.warning(f"{len(missing_div_teams)} team(s) missing a division — open 'Team divisions + opening balance' to fix.")
    else:
        st.caption("Team divisions/opening balance: OK (hidden below).")

    # ------------------------------------------------------------
    # 1) Team Divisions and Points Ladder
    # ------------------------------------------------------------
    st.markdown("### 1) Team Divisions and Points Ladder")

    with st.expander(
        "Edit team setup (only required once)",
        expanded=bool(missing_div_teams),
    ):

        st.write("Set division + opening balance per team (saved immediately).")

        div_col1, div_col2 = st.columns([2, 1], gap="large")

        with div_col1:
            for t in teams_today:
                cur_div = (existing.get(t, {}).get("division") or "").strip()
                cur_open = int(existing.get(t, {}).get("opening") or 0)

                default_div_idx = DIVISIONS.index(cur_div) if cur_div in DIVISIONS else 0

                r1, r2 = st.columns([2, 1], gap="medium")
                with r1:
                    new_div = st.selectbox(
                        label=t,
                        options=DIVISIONS,
                        index=default_div_idx,
                        key=f"div_select_{selected_date.isoformat()}_{t}",
                    )
                with r2:
                    new_open = st.number_input(
                        label="Opening",
                        min_value=0,
                        step=1,
                        value=cur_open,
                        key=f"open_{selected_date.isoformat()}_{t}",
                    )

                if new_div != cur_div or int(new_open) != cur_open:
                    upsert_team(t, new_div, int(new_open))
                    existing[t] = {"division": new_div, "opening": int(new_open)}

        with div_col2:
            st.write("Teams (today)")

            df_teams_today = pd.DataFrame(
                [
                    {
                        "Team": t,
                        "Division": (existing.get(t, {}).get("division") or "").strip() or "—",
                        "Opening": int(existing.get(t, {}).get("opening") or 0),
                    }
                    for t in teams_today
                ]
            )

            if df_teams_today.empty:
                st.caption("No teams found.")
            else:
                df_teams_today = df_teams_today.sort_values(
                    by=["Division", "Opening", "Team"],
                    ascending=[True, False, True],
                ).reset_index(drop=True)

                row_h = 35
                height = min(900, (len(df_teams_today) + 1) * row_h + 10)

                st.dataframe(
                    df_teams_today,
                    use_container_width=True,
                    hide_index=True,
                    height=height,
                )

    st.markdown("---")
    st.markdown("### 2) Enter game results (scores + referee inputs)")

    if not todays_games:
        st.info("No games found for this date.")
        st.stop()

    for g in todays_games:
        start_dt = dtparser.parse(g["start_dt"])

        # Fetch saved result (if any) for this game
        gr = get_game_result(int(g["id"]))
        is_saved = bool(gr)

        # Expander label (HTML won't reliably render in the label, so use an icon)
        icon = "✅" if is_saved else "⬜"
        label = f"{icon} {g['home_team']} vs {g['away_team']} — {g['field_name']} @ {_time_12h(start_dt)}"

        with st.expander(label, expanded=False):
            # Defaults (pulled from saved result if it exists)
            d_home_score = int(gr["home_score"]) if gr else 0
            d_away_score = int(gr["away_score"]) if gr else 0

            d_hft = int(gr["home_female_tries"]) if gr else 0
            d_aft = int(gr["away_female_tries"]) if gr else 0

            DEFAULT_CONDUCT = 10
            d_hc = int(gr["home_conduct"]) if gr else DEFAULT_CONDUCT
            d_ac = int(gr["away_conduct"]) if gr else DEFAULT_CONDUCT

            d_hu = int(gr["home_unstripped"]) if gr else 0
            d_au = int(gr["away_unstripped"]) if gr else 0

            d_hd = int(gr["home_defaulted"]) if gr else 0
            d_ad = int(gr["away_defaulted"]) if gr else 0

            left_col, right_col = st.columns([1.25, 1.0], gap="large")

            # ============================================================
            # LEFT: scoring inputs (per game)
            # ============================================================
            with left_col:
                col_team, col_score, col_ft, col_conduct, col_un, col_msg = st.columns(
                    [1.55, 0.85, 1.05, 0.95, 1.05, 2.55],
                    gap="small",
                )
                col_team.markdown("**Team**")
                col_score.markdown("**Score**")
                col_ft.markdown("**Female**")
                col_conduct.markdown("**Conduct**")
                col_un.markdown("**Unstrip**")
                col_msg.markdown("**Default / Instructions**")

                r1_team, r1_score, r1_ft, r1_conduct, r1_un, r1_msg = st.columns(
                    [1.55, 0.85, 1.05, 0.95, 1.05, 2.55],
                    gap="small",
                )
                r2_team, r2_score, r2_ft, r2_conduct, r2_un, r2_msg = st.columns(
                    [1.55, 0.85, 1.05, 0.95, 1.05, 2.55],
                    gap="small",
                )

                r1_team.markdown(f"**{g['home_team']}**")
                r2_team.markdown(f"**{g['away_team']}**")

                home_score = r1_score.number_input(
                    "home_score",
                    min_value=0,
                    step=1,
                    value=int(d_home_score),
                    key=f"hs_{g['id']}",
                    label_visibility="collapsed",
                )
                away_score = r2_score.number_input(
                    "away_score",
                    min_value=0,
                    step=1,
                    value=int(d_away_score),
                    key=f"as_{g['id']}",
                    label_visibility="collapsed",
                )

                home_ft = r1_ft.number_input(
                    "home_ft",
                    min_value=0,
                    step=1,
                    value=int(d_hft),
                    key=f"hft_{g['id']}",
                    label_visibility="collapsed",
                )
                away_ft = r2_ft.number_input(
                    "away_ft",
                    min_value=0,
                    step=1,
                    value=int(d_aft),
                    key=f"aft_{g['id']}",
                    label_visibility="collapsed",
                )

                home_conduct = r1_conduct.number_input(
                    "home_conduct",
                    min_value=0,
                    max_value=10,
                    step=1,
                    value=int(d_hc),
                    key=f"hc_{g['id']}",
                    label_visibility="collapsed",
                )
                away_conduct = r2_conduct.number_input(
                    "away_conduct",
                    min_value=0,
                    max_value=10,
                    step=1,
                    value=int(d_ac),
                    key=f"ac_{g['id']}",
                    label_visibility="collapsed",
                )

                home_un = r1_un.number_input(
                    "home_un",
                    min_value=0,
                    step=1,
                    value=int(d_hu),
                    key=f"hu_{g['id']}",
                    label_visibility="collapsed",
                )
                away_un = r2_un.number_input(
                    "away_un",
                    min_value=0,
                    step=1,
                    value=int(d_au),
                    key=f"au_{g['id']}",
                    label_visibility="collapsed",
                )

                home_defaulted = r1_msg.checkbox("DEFAULTED", value=bool(d_hd), key=f"hd_{g['id']}")
                away_defaulted = r2_msg.checkbox("DEFAULTED", value=bool(d_ad), key=f"ad_{g['id']}")

                if home_defaulted or away_defaulted:
                    if home_defaulted and away_defaulted:
                        r1_msg.markdown("⚠️ **Only ONE team can be marked DEFAULTED**")
                        r2_msg.markdown("⚠️ **Only ONE team can be marked DEFAULTED**")
                    elif home_defaulted:
                        r1_msg.markdown("⚠️ **Allocate 10 Conduct**")
                        r2_msg.markdown("**Allocate 10 Conduct + 3 Win**")
                    elif away_defaulted:
                        r2_msg.markdown("⚠️ **Allocate 10 Conduct**")
                        r1_msg.markdown("**Allocate 10 Conduct + 3 Win**")

                st.markdown("")

                if st.button("Save result", key=f"save_res_{g['id']}_ladder"):
                    if home_defaulted and away_defaulted:
                        st.error("Cannot save: only one team can be marked as DEFAULTED.")
                    else:
                        upsert_game_result(
                            game_id=int(g["id"]),
                            home_score=int(home_score),
                            away_score=int(away_score),
                            home_female_tries=int(home_ft),
                            away_female_tries=int(away_ft),
                            home_conduct=int(home_conduct),
                            away_conduct=int(away_conduct),
                            home_unstripped=int(home_un),
                            away_unstripped=int(away_un),
                            home_defaulted=1 if home_defaulted else 0,
                            away_defaulted=1 if away_defaulted else 0,
                        )
                        invalidate_ladder()
                        st.success("Saved.")
                        st.rerun()

            # ============================================================
            # RIGHT: live points breakdown (per game)
            # ============================================================
            with right_col:
                bd_live = compute_points_breakdown_for_game(
                    home_score=int(home_score),
                    away_score=int(away_score),
                    home_female_tries=int(home_ft),
                    away_female_tries=int(away_ft),
                    home_conduct=int(home_conduct),
                    away_conduct=int(away_conduct),
                    home_unstripped=int(home_un),
                    away_unstripped=int(away_un),
                    home_defaulted=1 if home_defaulted else 0,
                    away_defaulted=1 if away_defaulted else 0,
                )

                st.markdown("**Live points breakdown:**")

                def fmt_team(team: str, bd: dict) -> str:
                    return f"⚠️ {team}" if bool(bd.get("Defaulted")) else team

                h = bd_live["HOME"]
                a = bd_live["AWAY"]

                rows = [
                    {
                        "Team": fmt_team(g["home_team"], h),
                        "Score": f"{int(home_score)}",
                        "Match": f"{int(h.get('Match', 0))}",
                        "Female": f"{int(h.get('FemTries', 0))}",
                        "FemBP": f"{int(h.get('FemBP', 0))}",
                        "CloseBP": f"{int(h.get('CloseBP', 0))}",
                        "Conduct": f"{int(h.get('Conduct', 0))}",
                        "Pen": f"{int(h.get('Pen', 0))}",
                        "Unstrip": f"{int(h.get('Unstrip', 0))}",
                        "Res": f"{str(h.get('Res', '') or '')}",
                        "Total": f"{int(h.get('Points', 0))}",
                    },
                    {
                        "Team": fmt_team(g["away_team"], a),
                        "Score": f"{int(away_score)}",
                        "Match": f"{int(a.get('Match', 0))}",
                        "Female": f"{int(a.get('FemTries', 0))}",
                        "FemBP": f"{int(a.get('FemBP', 0))}",
                        "CloseBP": f"{int(a.get('CloseBP', 0))}",
                        "Conduct": f"{int(a.get('Conduct', 0))}",
                        "Pen": f"{int(a.get('Pen', 0))}",
                        "Unstrip": f"{int(a.get('Unstrip', 0))}",
                        "Res": f"{str(a.get('Res', '') or '')}",
                        "Total": f"{int(a.get('Points', 0))}",
                    },
                ]

                headers = ["Team", "Score", "Match", "Female", "FemBP", "CloseBP", "Conduct", "Pen", "Unstrip", "Res", "Total"]

                st.markdown(
                    """
                    <style>
                    .pts-grid {
                        display: grid;
                        grid-template-columns: 2.4fr 0.8fr 0.9fr 0.9fr 0.9fr 0.9fr 1.0fr 0.8fr 0.9fr 0.7fr 0.8fr;
                        gap: 6px 10px;
                        font-size: 12px;
                        line-height: 1.25;
                        white-space: nowrap;
                        align-items: center;
                    }
                    .pts-h {
                        font-weight: 700;
                        color: #444;
                        padding-bottom: 4px;
                        border-bottom: 1px solid #eee;
                    }
                    .pts-c { padding-top: 6px; }
                    .pts-total { font-weight: 800; }
                    </style>
                    """,
                    unsafe_allow_html=True,
                )

                header_html = "".join([f"<div class='pts-h'>{hh}</div>" for hh in headers])

                def row_html(r: dict) -> str:
                    return "".join(
                        [
                            f"<div class='pts-c'>{r['Team']}</div>",
                            f"<div class='pts-c'>{r['Score']}</div>",
                            f"<div class='pts-c'>{r['Match']}</div>",
                            f"<div class='pts-c'>{r['Female']}</div>",
                            f"<div class='pts-c'>{r['FemBP']}</div>",
                            f"<div class='pts-c'>{r['CloseBP']}</div>",
                            f"<div class='pts-c'>{r['Conduct']}</div>",
                            f"<div class='pts-c'>{r['Pen']}</div>",
                            f"<div class='pts-c'>{r['Unstrip']}</div>",
                            f"<div class='pts-c'>{r['Res']}</div>",
                            f"<div class='pts-c pts-total'>{r['Total']}</div>",
                        ]
                    )

                body_html = row_html(rows[0]) + row_html(rows[1])

                st.markdown(
                    f"<div class='pts-grid'>{header_html}{body_html}</div>",
                    unsafe_allow_html=True,
                )

    st.markdown("---")
    st.markdown("### 3) Ladder table (as at selected date)")

    warnings = ladder_validation_warnings_as_at(selected_date)
    if warnings:
        st.warning("Fix these issues to ensure ladder is correct:")
        for w in warnings:
            st.write(f"- {w}")

    all_divs = sorted({(r["division"] or "").strip() for r in list_teams()} - {""})
    if not all_divs:
        st.info("No divisions set yet. Assign divisions above first.")
        st.stop()

    division = st.selectbox(
        "Division",
        all_divs,
        key=f"ladder_div_{selected_date.isoformat()}",
    )

    _ = st.session_state.get("ladder_nonce", "0")   # ✅ makes ladder depend on the nonce
    df_ladder = ladder_table_df_as_at(selected_date, division)

    if df_ladder.empty:
        st.info("No ladder data yet for this division/date. Save some results first.")
    else:
        st.dataframe(df_ladder, use_container_width=True, hide_index=True)


# ============================================================
# Import tab
# ============================================================
with tabs[2]:
    st.subheader("Import CSVs")

    # ----------------------------
    # Games CSV
    # ----------------------------
    st.markdown("### Games CSV")
    st.caption("Required columns: game_id, date, start_time, home_team, away_team, field")

    replace_games_mode = st.checkbox(
        "Replace ALL games with this CSV (overwrite existing draw)",
        value=False,
        key="replace_games_mode",
    )

    games_file = st.file_uploader("Upload Games CSV", type=["csv"], key="games_csv")
    if games_file:
        df_games = pd.read_csv(games_file)
        st.dataframe(df_games.head(20), use_container_width=True)

        c1, c2 = st.columns([1, 2])
        with c1:
            if st.button("Import Games", key="import_games_btn"):
                try:
                    if replace_games_mode:
                        imported = replace_games_csv(df_games)
                        st.success(
                            f"✅ Replaced the entire draw. Imported {imported} game(s). "
                            "All assignments/offers/results were cleared."
                        )

                        # Clear UI selections that might reference old game IDs/dates
                        for k in [
                            "games_date_select",
                            "ladder_date_select",
                            "admin_summary_pdf_bytes",
                            "admin_summary_xlsx_bytes",
                            "ref_scorecards_pdf_bytes",
                        ]:
                            st.session_state.pop(k, None)

                        # Also clear any per-slot selectboxes
                        for k in [k for k in st.session_state.keys() if str(k).startswith("refpick_")]:
                            st.session_state.pop(k, None)

                    else:
                        ins, upd = import_games_csv(df_games)
                        st.success(f"Imported games. Inserted: {ins}, Updated: {upd}")

                    st.rerun()
                except Exception as e:
                    st.error(str(e))

        with c2:
            if replace_games_mode:
                st.warning(
                    "Replace mode will DELETE ALL existing games, assignments, offers, and ladder results "
                    "before importing this CSV."
                )

    st.markdown("---")

    # ----------------------------
    # Referees CSV
    # ----------------------------
    st.markdown("### Referees CSV")
    st.caption("Required columns: name, email (optional: phone)")

    replace_mode = st.checkbox(
        "Replace ALL referees with this CSV (resets assignments + offers + blackouts)",
        value=False,
        key="replace_refs_mode",
    )

    refs_file = st.file_uploader("Upload Referees CSV", type=["csv"], key="refs_csv")
    if refs_file:
        df_refs = pd.read_csv(refs_file)
        st.dataframe(df_refs.head(20), use_container_width=True)

        if st.button("Import Referees", key="import_refs_btn"):
            try:
                if replace_mode:
                    count = replace_referees_csv(df_refs)
                    st.success(
                        f"✅ Replaced referee list successfully. Imported {count} referee(s). "
                        "All assignments were reset to EMPTY."
                    )
                    for k in [k for k in st.session_state.keys() if str(k).startswith("refpick_")]:
                        st.session_state.pop(k, None)
                else:
                    added, updated = import_referees_csv(df_refs)
                    st.success(f"Imported referees. Added: {added}, Updated: {updated}")

                st.rerun()
            except Exception as e:
                st.error(str(e))

    st.markdown("---")

    # ----------------------------
    # Blackouts CSV (optional)
    # ----------------------------
    st.markdown("### Blackouts CSV (optional)")
    st.caption("Required columns: email, blackout_date")

    bl_file = st.file_uploader("Upload Blackouts CSV", type=["csv"], key="bl_csv")
    if bl_file:
        df_bl = pd.read_csv(bl_file)
        st.dataframe(df_bl.head(20), use_container_width=True)

        if st.button("Import Blackouts", key="import_bl_btn"):
            try:
                added, skipped = import_blackouts_csv(df_bl)
                st.success(f"Imported blackouts. Added: {added}. Skipped (unknown referee): {skipped}")
                st.rerun()
            except Exception as e:
                st.error(str(e))

    st.markdown("---")
    st.subheader("Backups (DB snapshot)")
    st.caption("Backups are consistent snapshots of the entire league database (games, allocations, ladder, admins, etc).")

    c_b1, c_b2 = st.columns([1, 2], gap="large")

    with c_b1:
        backup_label = st.text_input("Optional label", value="", key="backup_label")
        if st.button("Create backup now", key="create_backup_now_btn"):
            try:
                p = create_backup_now(label=backup_label)
                st.success(f"Backup created: {p.name}")
            except Exception as e:
                st.error(f"Backup failed: {e}")

        last_b = meta_get("last_backup_at", "")
        last_r = meta_get("last_restore_at", "")
        st.caption(f"Last backup: {last_b or '—'}")
        st.caption(f"Last restore: {last_r or '—'}")

    with c_b2:
        backups = list_backups()
        if not backups:
            st.info("No backups found yet.")
        else:
            options = [p.name for p in backups]
            pick_name = st.selectbox("Select a backup", options, key="backup_pick")
            pick_path = next((p for p in backups if p.name == pick_name), None)

            d1, d2, d3 = st.columns([1, 1, 2])

            with d1:
                if pick_path and pick_path.exists():
                    st.download_button(
                        "Download selected",
                        data=pick_path.read_bytes(),
                        file_name=pick_path.name,
                        mime="application/octet-stream",
                        key="download_backup_btn",
                    )

            with d2:
                st.caption(f"Size: {(pick_path.stat().st_size/1024/1024):.2f} MB" if pick_path else "")

            with d3:
                st.warning("Restore will overwrite the CURRENT database.")
                confirm1 = st.checkbox("I understand this overwrites current data", key="restore_confirm_1")
                confirm2 = st.checkbox("Yes — restore selected backup now", key="restore_confirm_2")

                if st.button("RESTORE selected backup", key="restore_selected_btn", disabled=not (confirm1 and confirm2)):
                    if not pick_path or not pick_path.exists():
                        st.error("Selected backup file could not be found.")
                    else:
                        try:
                            pre = create_backup_now(label="pre_restore")
                            restore_from_backup_file(pick_path)
                            st.success(f"Restored from {pick_path.name}. (Pre-restore backup: {pre.name})")
                            st.rerun()
                        except Exception as e:
                            st.error(f"Restore failed: {e}")

        st.markdown("#### Restore from uploaded backup (.db)")
        up = st.file_uploader("Upload a backup DB file", type=["db"], key="upload_backup_db")
        confirm_up = st.checkbox("I understand upload-restore overwrites current data", key="restore_upload_confirm")
        if up and confirm_up:
            if st.button("RESTORE uploaded backup", key="restore_uploaded_btn"):
                try:
                    tmp_path = BACKUPS_DIR / f"uploaded_restore_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.db"
                    tmp_path.write_bytes(up.read())

                    pre = create_backup_now(label="pre_restore_upload")
                    restore_from_backup_file(tmp_path)

                    st.success(f"Upload restored. (Pre-restore backup: {pre.name})")
                    st.rerun()
                except Exception as e:
                    st.error(f"Upload restore failed: {e}")

# ============================================================
# Blackouts tab
# ============================================================
with tabs[3]:
    st.subheader("Manage Blackout Dates (date-only)")

    refs = get_referees()
    if not refs:
        st.info("Import referees first.")
        st.stop()

    left_col, right_col = st.columns([2, 1], gap="large")

    # ----------------------------
    # LEFT: existing add/remove UI
    # ----------------------------
    with left_col:
        ref_map = {f"{r['name']} ({r['email']})": r["id"] for r in refs}
        choice = st.selectbox("Select referee", list(ref_map.keys()), key="blackout_ref_select")
        ref_id = ref_map[choice]

        add_date = st.date_input("Add blackout date", value=date.today(), key="blackout_add_date")
        if st.button("Add date", key="blackout_add_btn"):
            conn = db()
            try:
                conn.execute(
                    "INSERT INTO blackouts(referee_id, blackout_date) VALUES (?, ?)",
                    (ref_id, add_date.isoformat()),
                )
                conn.commit()
                st.success("Added blackout date.")
                st.rerun()
            except sqlite3.IntegrityError:
                st.warning("That date is already in the blackout list.")
            finally:
                conn.close()

        st.markdown("### Current blackout dates (selected referee)")
        conn = db()
        rows = conn.execute(
            """
            SELECT blackout_date FROM blackouts
            WHERE referee_id=?
            ORDER BY blackout_date ASC
            """,
            (ref_id,),
        ).fetchall()
        conn.close()

        if rows:
            dates = [r["blackout_date"] for r in rows]
            del_date = st.selectbox("Remove blackout date", dates, key="blackout_del_select")
            if st.button("Remove selected date", key="blackout_del_btn"):
                conn = db()
                conn.execute(
                    "DELETE FROM blackouts WHERE referee_id=? AND blackout_date=?",
                    (ref_id, del_date),
                )
                conn.commit()
                conn.close()
                st.success("Removed.")
                st.rerun()
        else:
            st.caption("No blackout dates set for this referee.")

    # ----------------------------
    # RIGHT: ALL blackouts table
    # ----------------------------
    with right_col:
        st.markdown("### All blackout dates")
        st.caption("Sorted by date (blank row between dates)")

        df_all = list_all_blackouts_with_ref_names()

        if df_all.empty:
            st.caption("No blackout dates in the system.")
        else:
            row_h = 32
            height = min(900, (len(df_all) + 1) * row_h + 10)

            st.dataframe(
                df_all,
                use_container_width=True,
                hide_index=True,
                height=height,
            )


# ============================================================
# Administrators tab
# ============================================================
with tabs[4]:
    st.subheader("Administrators (allowlist)")
    st.caption("Add/remove admins by email. Removed admins lose access immediately.")

    # Only Super Admin can manage administrators
    if not is_super_admin_logged_in():
        st.info("Only the Super Admin can manage administrators.")
        st.stop()

    admins = list_admins()
    if admins:
        df_admins = pd.DataFrame(
            [
                {
                    "email": a["email"],
                    "active": "YES" if a["active"] == 1 else "NO",
                    "created_at": a["created_at"],
                }
                for a in admins
            ]
        )
        st.dataframe(df_admins, use_container_width=True)

    st.markdown("### Add admin")
    new_admin = st.text_input("Email to add", key="add_admin_email")
    if st.button("Add admin", key="add_admin_btn"):
        if not new_admin.strip():
            st.error("Enter an email.")
        else:
            add_admin(new_admin)
            st.success("Admin added (or already existed).")
            st.rerun()

    st.markdown("### Remove/disable admin")
    active_emails = [a["email"] for a in admins if a["active"] == 1]
    if active_emails:
        disable_email = st.selectbox("Select admin to disable", active_emails, key="disable_admin_select")
        if st.button("Disable selected admin", key="disable_admin_btn"):
            if disable_email == st.session_state.get("admin_email"):
                st.error("You can't disable yourself while logged in.")
            else:
                set_admin_active(disable_email, False)
                st.success("Admin disabled.")
                st.rerun()
    else:
        st.info("No active admins found (you should add at least one).")